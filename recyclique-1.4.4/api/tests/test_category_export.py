"""Tests for category export functionality (Story B21-P3)."""

import pytest
from io import BytesIO
from decimal import Decimal
from uuid import uuid4

from openpyxl import load_workbook
from PyPDF2 import PdfReader

from recyclic_api.models.category import Category
from recyclic_api.models.user import User, UserRole, UserStatus
from recyclic_api.core.security import hash_password


class TestCategoryExportEndpoint:
    """Test export endpoint with authentication and authorization"""

    def test_export_pdf_requires_super_admin_role(self, client, db_session):
        """Test that PDF export requires SUPER_ADMIN role"""
        # Create a regular user
        regular_user = User(
            id=uuid4(),
            username="regular@test.com",
            hashed_password=hash_password("testpassword"),
            role=UserRole.USER,
            status=UserStatus.ACTIVE
        )
        db_session.add(regular_user)
        db_session.commit()

        # Try to export with regular user - should fail with 403
        from recyclic_api.core.security import create_access_token
        token = create_access_token(data={"sub": str(regular_user.id)})
        client.headers = {"Authorization": f"Bearer {token}"}

        response = client.get("/api/v1/categories/actions/export?format=pdf")
        assert response.status_code == 403

    def test_export_xls_requires_super_admin_role(self, client, db_session):
        """Test that XLS export requires SUPER_ADMIN role"""
        # Create a regular user
        regular_user = User(
            id=uuid4(),
            username="regular2@test.com",
            hashed_password=hash_password("testpassword"),
            role=UserRole.USER,
            status=UserStatus.ACTIVE
        )
        db_session.add(regular_user)
        db_session.commit()

        # Try to export with regular user - should fail with 403
        from recyclic_api.core.security import create_access_token
        token = create_access_token(data={"sub": str(regular_user.id)})
        client.headers = {"Authorization": f"Bearer {token}"}

        response = client.get("/api/v1/categories/actions/export?format=xls")
        assert response.status_code == 403

    def test_export_pdf_success(self, super_admin_client, db_session):
        """Test successful PDF export with SUPER_ADMIN role"""
        # Create test categories
        cat1 = Category(
            id=uuid4(),
            name="Electronics",
            is_active=True,
            price=Decimal("10.00"),
            max_price=Decimal("50.00")
        )
        cat2 = Category(
            id=uuid4(),
            name="Furniture",
            is_active=True,
            price=Decimal("20.00"),
            max_price=Decimal("100.00")
        )
        db_session.add_all([cat1, cat2])
        db_session.commit()

        # Export PDF
        response = super_admin_client.get("/api/v1/categories/actions/export?format=pdf")

        # Assert response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "attachment" in response.headers["content-disposition"]
        assert "categories_export_" in response.headers["content-disposition"]
        assert ".pdf" in response.headers["content-disposition"]

        # Verify PDF content
        pdf_bytes = BytesIO(response.content)
        pdf_reader = PdfReader(pdf_bytes)
        assert len(pdf_reader.pages) > 0

        # Extract text from first page
        first_page_text = pdf_reader.pages[0].extract_text()
        assert "Configuration des Catégories" in first_page_text

    def test_export_xls_success(self, super_admin_client, db_session):
        """Test successful Excel export with SUPER_ADMIN role"""
        # Create test categories with hierarchy
        root_cat = Category(
            id=uuid4(),
            name="Root Category",
            is_active=True
        )
        db_session.add(root_cat)
        db_session.commit()

        child_cat = Category(
            id=uuid4(),
            name="Child Category",
            is_active=True,
            parent_id=root_cat.id,
            price=Decimal("15.00"),
            max_price=Decimal("75.00")
        )
        db_session.add(child_cat)
        db_session.commit()

        # Export Excel
        response = super_admin_client.get("/api/v1/categories/actions/export?format=xls")

        # Assert response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert "categories_export_" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

        # Verify Excel content
        excel_bytes = BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        ws = wb.active

        # Check headers
        assert ws['A1'].value == "ID Catégorie Parente"
        assert ws['B1'].value == "Nom Catégorie"
        assert ws['C1'].value == "Prix Minimum"
        assert ws['D1'].value == "Prix Maximum"

        # Check data rows (at least 2: root + child)
        assert ws.max_row >= 3  # Header + at least 2 data rows

        # Verify root category is in the export
        category_names = [ws[f'B{i}'].value for i in range(2, ws.max_row + 1)]
        assert "Root Category" in category_names
        assert "Child Category" in category_names

    def test_export_invalid_format(self, super_admin_client, db_session):
        """Test export with invalid format parameter"""
        response = super_admin_client.get("/api/v1/categories/actions/export?format=invalid")

        assert response.status_code == 400
        assert "Invalid format" in response.json()["detail"]

    def test_export_missing_format_parameter(self, super_admin_client, db_session):
        """Test export without format parameter"""
        response = super_admin_client.get("/api/v1/categories/actions/export")

        assert response.status_code == 422  # FastAPI validation error

#    def test_export_pdf_with_empty_database(self, super_admin_client, db_session):
#        """Test PDF export when no categories exist"""
#        # Delete all categories if any
#        db_session.query(Category).delete()
#        db_session.commit()
#
#        # Export should still work, just with empty content
#        response = super_admin_client.get("/api/v1/categories/actions/export?format=pdf")
#
#        assert response.status_code == 200
#        assert response.headers["content-type"] == "application/pdf"
#
#    def test_export_xls_with_empty_database(self, super_admin_client, db_session):
#        """Test Excel export when no categories exist"""
#        # Delete all categories if any
#        db_session.query(Category).delete()
#        db_session.commit()
#
#        # Export should still work, just with headers only
#        response = super_admin_client.get("/api/v1/categories/actions/export?format=xls")
#
#        assert response.status_code == 200
#
#        # Verify Excel has headers but no data rows
#        excel_bytes = BytesIO(response.content)
#        wb = load_workbook(excel_bytes)
#        ws = wb.active
#
#        # Should have header row
#        assert ws['A1'].value == "ID Catégorie Parente"
#        # May have only header row (max_row == 1) or be 2 if there's an empty row
#        assert ws.max_row <= 2

    def test_export_pdf_with_hierarchy(self, super_admin_client, db_session):
        """Test PDF export preserves category hierarchy"""
        # Create hierarchy: Parent -> Child -> Grandchild
        parent = Category(
            id=uuid4(),
            name="Parent Category",
            is_active=True
        )
        db_session.add(parent)
        db_session.commit()

        child = Category(
            id=uuid4(),
            name="Child Category",
            is_active=True,
            parent_id=parent.id
        )
        db_session.add(child)
        db_session.commit()

        grandchild = Category(
            id=uuid4(),
            name="Grandchild Category",
            is_active=True,
            parent_id=child.id,
            price=Decimal("25.00")
        )
        db_session.add(grandchild)
        db_session.commit()

        # Export PDF
        response = super_admin_client.get("/api/v1/categories/actions/export?format=pdf")

        assert response.status_code == 200

        # Verify PDF contains all categories
        pdf_bytes = BytesIO(response.content)
        pdf_reader = PdfReader(pdf_bytes)
        full_text = ""
        for page in pdf_reader.pages:
            full_text += page.extract_text()

        assert "Parent Category" in full_text
        assert "Child Category" in full_text
        assert "Grandchild Category" in full_text

    def test_export_xls_with_prices(self, super_admin_client, db_session):
        """Test Excel export includes price information"""
        cat = Category(
            id=uuid4(),
            name="Priced Category",
            is_active=True,
            price=Decimal("12.50"),
            max_price=Decimal("99.99")
        )
        db_session.add(cat)
        db_session.commit()

        # Export Excel
        response = super_admin_client.get("/api/v1/categories/actions/export?format=xls")

        assert response.status_code == 200

        # Verify prices in Excel
        excel_bytes = BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        ws = wb.active

        # Find the row with "Priced Category"
        for row in range(2, ws.max_row + 1):
            if ws[f'B{row}'].value == "Priced Category":
                assert ws[f'C{row}'].value == 12.50
                assert ws[f'D{row}'].value == 99.99
                break
        else:
            pytest.fail("Priced Category not found in export")


@pytest.fixture
def super_admin_client(client, db_session):
    """Fixture to create a client authenticated as SUPER_ADMIN"""
    # Create super admin user
    super_admin = User(
        id=uuid4(),
        username="superadmin@test.com",
        hashed_password=hash_password("adminpassword"),
        role=UserRole.SUPER_ADMIN,
        status=UserStatus.ACTIVE
    )
    db_session.add(super_admin)
    db_session.commit()

    # Create access token
    from recyclic_api.core.security import create_access_token
    token = create_access_token(data={"sub": str(super_admin.id)})

    # Set authorization header
    client.headers = {"Authorization": f"Bearer {token}"}

    return client
