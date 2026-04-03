"""Service for exporting categories to PDF and Excel formats (Story B21-P3)."""

from io import BytesIO
import io
import csv
from datetime import datetime
from typing import List, Optional, Tuple
from decimal import Decimal

from sqlalchemy.orm import Session
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..models.category import Category
from ..schemas.category import CategoryRead


class CategoryExportService:
    """Service for exporting categories in PDF and Excel formats"""

    def __init__(self, db: Session):
        self.db = db

    def _get_all_categories_hierarchy(self) -> List[Tuple[Category, int]]:
        """Get all categories ordered hierarchically (root first, then children)"""
        # Get all active categories
        categories = self.db.query(Category).filter(Category.is_active == True).order_by(Category.name).all()

        # Build hierarchy: root categories first, then their children recursively
        hierarchy = []
        root_categories = [cat for cat in categories if cat.parent_id is None]

        def add_category_and_children(category: Category, level: int = 0):
            hierarchy.append((category, level))
            children = [cat for cat in categories if cat.parent_id == category.id]
            for child in sorted(children, key=lambda x: x.name):
                add_category_and_children(child, level + 1)

        for root in sorted(root_categories, key=lambda x: x.name):
            add_category_and_children(root)

        return hierarchy

    def _format_price(self, price: Optional[Decimal]) -> str:
        """Format price for display"""
        if price is None:
            return "-"
        return f"{float(price):.2f} €"

    def export_to_pdf(self) -> BytesIO:
        """
        Generate a PDF export of all categories with professional layout.
        Each root category is displayed with its name as a title, followed by its children in a table.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=2*cm,
            rightMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#7f8c8d'),
            spaceAfter=20,
            alignment=TA_CENTER
        )

        category_title_style = ParagraphStyle(
            'CategoryTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=10,
            spaceBefore=10,
            alignment=TA_CENTER,
            borderWidth=2,
            borderColor=colors.HexColor('#3498db'),
            borderPadding=8
        )

        # Main title
        elements.append(Paragraph("Configuration des Catégories", title_style))
        elements.append(Paragraph(
            f"Export généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            subtitle_style
        ))
        elements.append(Spacer(1, 0.5*cm))

        # Get categories hierarchy
        hierarchy = self._get_all_categories_hierarchy()

        if not hierarchy:
            elements.append(Paragraph("Aucune catégorie active trouvée.", styles['Normal']))
        else:
            # Process categories grouped by root category
            root_categories = [cat for cat, level in hierarchy if level == 0]

            for root_idx, root_cat in enumerate(root_categories):
                # Build a group block (title + spacer + table) and keep it together across pages
                group_block = []
                group_block.append(Paragraph(f"{root_cat.name}", category_title_style))
                group_block.append(Spacer(1, 0.3*cm))

                # Collect children for this root
                group_data = []

                # Add header row for the table
                group_data.append([
                    Paragraph('<b>Nom Sous-Catégorie</b>', styles['Normal']),
                    Paragraph('<b>Prix Min</b>', styles['Normal']),
                    Paragraph('<b>Prix Max</b>', styles['Normal'])
                ])

                # Add root category itself if it has prices
                if root_cat.price is not None or root_cat.max_price is not None:
                    group_data.append([
                        Paragraph(root_cat.name, styles['Normal']),
                        Paragraph(self._format_price(root_cat.price), styles['Normal']),
                        Paragraph(self._format_price(root_cat.max_price), styles['Normal'])
                    ])
                else:
                    # Add empty row if root has no prices
                    group_data.append([
                        Paragraph("", styles['Normal']),
                        Paragraph("", styles['Normal']),
                        Paragraph("", styles['Normal'])
                    ])

                # Add children
                for cat, level in hierarchy:
                    if level > 0 and self._is_descendant_of(cat, root_cat.id, hierarchy):
                        group_data.append([
                            Paragraph(cat.name, styles['Normal']),
                            Paragraph(self._format_price(cat.price), styles['Normal']),
                            Paragraph(self._format_price(cat.max_price), styles['Normal'])
                        ])

                # Create table for this group
                col_widths = [10*cm, 3*cm, 3*cm]
                table = Table(group_data, colWidths=col_widths)

                # Table styling
                table.setStyle(TableStyle([
                    # Header row
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

                    # Data rows
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2c3e50')),
                    ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.lightgrey]),

                    # Grid
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                # Add table to group and KeepTogether the whole block (title + spacer + table)
                group_block.append(table)
                elements.append(KeepTogether(group_block))

                # Add spacing between root categories (but not after the last one)
                if root_idx < len(root_categories) - 1:
                    elements.append(Spacer(1, 1*cm))

        # Footer
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#95a5a6'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(
            f"Document généré par RecyClique - {datetime.now().year}",
            footer_style
        ))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _is_descendant_of(self, cat: Category, root_id: str, hierarchy: List[Tuple[Category, int]]) -> bool:
        """Check if a category is a descendant of the given root category"""
        current = cat
        while current.parent_id is not None:
            if current.parent_id == root_id:
                return True
            # Find parent in hierarchy
            parent = next((c for c, _ in hierarchy if c.id == current.parent_id), None)
            if parent is None:
                break
            current = parent
        return False

    def export_to_excel(self) -> BytesIO:
        """
        Generate an Excel export of all categories.
        Structure: First column shows root category name, second column shows sub-category name.
        Returns a BytesIO buffer containing the .xlsx file.
        """
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Catégories"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = ["Catégorie Racine", "Nom Sous-Catégorie", "Prix Minimum", "Prix Maximum", "Info", "Image URL"]
        ws.append(headers)

        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Get categories hierarchy
        hierarchy = self._get_all_categories_hierarchy()

        # Build a map to find root category for each category
        def get_root_category(cat: Category) -> Category:
            """Get the root category for any category"""
            current = cat
            while current.parent_id is not None:
                parent = next((c for c, _ in hierarchy if c.id == current.parent_id), None)
                if parent is None:
                    break
                current = parent
            return current

        # Data rows
        for cat, level in hierarchy:
            root_cat = get_root_category(cat)

            # For root categories, show their name in first column and empty second column
            # For children, show root name in first column and child name in second column
            if level == 0:
                # This is a root category
                row = [
                    root_cat.name,
                    "",  # Empty for root category itself
                    float(cat.price) if cat.price is not None else "",
                    float(cat.max_price) if cat.max_price is not None else "",
                    "",  # Info column (empty for now)
                    ""   # Image URL column (empty for now)
                ]
            else:
                # This is a child category
                row = [
                    root_cat.name,
                    cat.name,
                    float(cat.price) if cat.price is not None else "",
                    float(cat.max_price) if cat.max_price is not None else "",
                    "",  # Info column (empty for now)
                    ""   # Image URL column (empty for now)
                ]
            ws.append(row)

        # Column widths
        ws.column_dimensions['A'].width = 30  # Catégorie Racine
        ws.column_dimensions['B'].width = 30  # Nom Sous-Catégorie
        ws.column_dimensions['C'].width = 15  # Prix Min
        ws.column_dimensions['D'].width = 15  # Prix Max
        ws.column_dimensions['E'].width = 20  # Info
        ws.column_dimensions['F'].width = 20  # Image URL

        # Cell alignment for data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[0].alignment = Alignment(horizontal="left")   # Root Category
            row[1].alignment = Alignment(horizontal="left")   # Sub-Category Name
            row[2].alignment = Alignment(horizontal="right")  # Min Price
            row[3].alignment = Alignment(horizontal="right")  # Max Price
            row[4].alignment = Alignment(horizontal="left")   # Info
            row[5].alignment = Alignment(horizontal="left")   # Image URL

        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_to_csv(self) -> bytes:
        """
        Generate a CSV export aligned with the import template headers so it can be re-imported.
        Headers: "Catégorie racine","Sous-catégorie","Prix minimum (€)","Prix maximum (€)"
        Only root and first-level children are exported to match the import contract.
        """
        hierarchy = self._get_all_categories_hierarchy()

        # Build a quick map from category to level
        level_map = {cat.id: level for cat, level in hierarchy}

        # Helper to find root for any category
        def get_root(cat: Category) -> Category:
            current = cat
            while current.parent_id is not None:
                parent = next((c for c, _ in hierarchy if c.id == current.parent_id), None)
                if parent is None:
                    break
                current = parent
            return current

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Catégorie racine", "Sous-catégorie", "Prix minimum (€)", "Prix maximum (€)"])

        for cat, level in hierarchy:
            # Root rows
            if level == 0:
                writer.writerow([
                    cat.name,
                    "",
                    f"{float(cat.price):.2f}" if cat.price is not None else "",
                    f"{float(cat.max_price):.2f}" if cat.max_price is not None else "",
                ])
            # First-level children only to respect import contract
            elif level == 1:
                root = get_root(cat)
                writer.writerow([
                    root.name,
                    cat.name,
                    f"{float(cat.price):.2f}" if cat.price is not None else "",
                    f"{float(cat.max_price):.2f}" if cat.max_price is not None else "",
                ])

        return buf.getvalue().encode("utf-8")
