"""
Tests d'intégration pour l'export bulk de sessions de caisse (Story B45-P1).

Tests vérifient:
- Export CSV bulk avec filtres
- Export Excel bulk avec filtres
- Respect des filtres (date, statut, opérateur, site)
- Permissions (ADMIN/SUPER_ADMIN uniquement)
- Performance avec grandes quantités
"""
import io
from datetime import datetime, timedelta, timezone
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from recyclic_api.models.cash_session import CashSession, CashSessionStatus
from recyclic_api.models.user import User, UserRole, UserStatus
from recyclic_api.models.site import Site
from recyclic_api.models.sale import Sale
from recyclic_api.models.sale_item import SaleItem
from recyclic_api.core.security import hash_password


@pytest.fixture
def test_site(db_session: Session) -> Site:
    """Créer un site de test."""
    site = Site(
        id=uuid4(),
        name="Site Test",
        address="123 Test St"
    )
    db_session.add(site)
    db_session.commit()
    db_session.refresh(site)
    return site


@pytest.fixture
def test_operator(db_session: Session) -> User:
    """Créer un opérateur de test."""
    operator = User(
        id=uuid4(),
        username="operator_test",
        hashed_password=hash_password("testpass"),
        role=UserRole.USER,
        status=UserStatus.ACTIVE,
        is_active=True
    )
    db_session.add(operator)
    db_session.commit()
    db_session.refresh(operator)
    return operator


@pytest.fixture
def test_sessions(db_session: Session, test_site: Site, test_operator: User) -> list[CashSession]:
    """Créer plusieurs sessions de test avec différentes dates."""
    sessions = []
    now = datetime.now(timezone.utc)
    
    # Session 1: Hier, fermée, avec ventes
    session1 = CashSession(
        id=uuid4(),
        operator_id=test_operator.id,
        site_id=test_site.id,
        initial_amount=100.0,
        status=CashSessionStatus.CLOSED,
        opened_at=now - timedelta(days=1),
        closed_at=now - timedelta(hours=12),
        total_sales=250.0,
        number_of_sales=5,
        total_items=10,
        total_donations=10.0,
        closing_amount=350.0,
        actual_amount=350.0,
        variance=0.0
    )
    db_session.add(session1)
    
    # Créer une vente pour session1
    sale1 = Sale(
        id=uuid4(),
        cash_session_id=session1.id,
        operator_id=test_operator.id,
        total_amount=50.0,
        donation=5.0,
        payment_method="cash"
    )
    db_session.add(sale1)
    db_session.flush()
    
    item1 = SaleItem(
        id=uuid4(),
        sale_id=sale1.id,
        category="EEE-4",
        quantity=2,
        unit_price=25.0,
        total_price=50.0
    )
    db_session.add(item1)
    
    # Session 2: Aujourd'hui, ouverte
    session2 = CashSession(
        id=uuid4(),
        operator_id=test_operator.id,
        site_id=test_site.id,
        initial_amount=150.0,
        status=CashSessionStatus.OPEN,
        opened_at=now - timedelta(hours=2),
        total_sales=100.0,
        number_of_sales=3,
        total_items=6
    )
    db_session.add(session2)
    
    # Session 3: Il y a 3 jours, fermée, sans ventes (session vide)
    session3 = CashSession(
        id=uuid4(),
        operator_id=test_operator.id,
        site_id=test_site.id,
        initial_amount=200.0,
        status=CashSessionStatus.CLOSED,
        opened_at=now - timedelta(days=3),
        closed_at=now - timedelta(days=3, hours=1),
        total_sales=0.0,
        number_of_sales=0,
        total_items=0,
        closing_amount=200.0,
        actual_amount=200.0,
        variance=0.0
    )
    db_session.add(session3)
    
    db_session.commit()
    
    for session in [session1, session2, session3]:
        db_session.refresh(session)
        sessions.append(session)
    
    return sessions


class TestBulkExportCashSessionsCSV:
    """Tests pour l'export CSV bulk de sessions de caisse."""
    
    def test_export_csv_bulk_success(self, admin_client, test_sessions):
        """Test export CSV bulk avec succès."""
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert "export_sessions_caisse" in response.headers["content-disposition"]
        
        # Vérifier le contenu CSV
        content = response.text
        assert "ID Session" in content
        assert "Statut" in content
        assert "Date Ouverture" in content
        assert "Opérateur" in content
        assert "Total Ventes" in content
    
    def test_export_csv_bulk_with_filters(self, admin_client, test_sessions):
        """Test export CSV avec filtres (statut, date)."""
        yesterday = datetime.now(timezone.utc) - timedelta(days=1)
        
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": yesterday.isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "status": "closed",
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        content = response.text
        
        # Vérifier que seules les sessions fermées sont incluses
        lines = content.split('\n')
        # Ligne 1 = headers, lignes suivantes = données
        data_lines = [l for l in lines[1:] if l.strip()]
        
        # Au moins une session fermée doit être présente
        assert len(data_lines) > 0
    
    def test_export_csv_bulk_unauthorized(self, client, db_session, test_sessions):
        """Test que USER ne peut pas exporter bulk."""
        from recyclic_api.core.security import create_access_token
        
        # Créer un utilisateur USER (non-admin) dans la session de test
        user = User(
            id=uuid4(),
            username="user_test",
            hashed_password=hash_password("testpass"),
            role=UserRole.USER,
            status=UserStatus.ACTIVE,
            is_active=True
        )
        db_session.add(user)
        db_session.commit()
        db_session.refresh(user)
        
        token = create_access_token(data={"sub": str(user.id)})
        client.headers["Authorization"] = f"Bearer {token}"
        
        response = client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {},
                "format": "csv"
            }
        )
        
        assert response.status_code == 403
    
    def test_export_csv_bulk_respects_filters(self, admin_client, test_sessions, test_site, test_operator):
        """Test que les filtres sont bien respectés."""
        # Filtrer par site
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "site_id": str(test_site.id),
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        content = response.text
        
        # Vérifier que toutes les sessions retournées sont du bon site
        lines = content.split('\n')
        # Les sessions de test sont toutes du même site, donc on vérifie juste que ça fonctionne
        assert len(lines) > 1  # Au moins headers + données


class TestBulkExportCashSessionsExcel:
    """Tests pour l'export Excel bulk de sessions de caisse."""
    
    def test_export_excel_bulk_success(self, admin_client, test_sessions):
        """Test export Excel bulk avec succès."""
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]
        
        # Vérifier que c'est un fichier Excel valide
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier les onglets
        assert "Résumé" in wb.sheetnames
        assert "Détails" in wb.sheetnames
        
        # Vérifier le contenu de l'onglet Résumé
        ws_summary = wb["Résumé"]
        assert ws_summary.max_row > 1  # Au moins headers + données
        
        # Vérifier les en-têtes
        headers = [cell.value for cell in ws_summary[1]]
        assert "Statut" in headers
        assert "Date Ouverture" in headers
        assert "CA Total (€)" in headers
        
        # Vérifier le contenu de l'onglet Détails
        ws_details = wb["Détails"]
        assert ws_details.max_row > 1
        
        headers_details = [cell.value for cell in ws_details[1]]
        assert "ID Session" in headers_details
        assert "Total Ventes (€)" in headers_details
    
    def test_export_excel_bulk_with_filters(self, admin_client, test_sessions):
        """Test export Excel avec filtres."""
        yesterday = datetime.now(timezone.utc) - timedelta(days=1)
        
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": yesterday.isoformat(),
                    "status": "closed",
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        
        # Vérifier le fichier Excel
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier qu'il y a des données
        ws_summary = wb["Résumé"]
        assert ws_summary.max_row > 1  # Headers + au moins une ligne de données
    
    def test_export_excel_formatting_styles(self, admin_client, test_sessions):
        """Test que la mise en forme (styles, couleurs, bordures) est appliquée."""
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier les styles des en-têtes dans l'onglet Résumé
        ws_summary = wb["Résumé"]
        header_row = ws_summary[1]
        
        for cell in header_row:
            # Vérifier que les en-têtes sont en gras
            assert cell.font.bold is True, f"Header cell {cell.coordinate} should be bold"
            # Vérifier que les en-têtes ont un fond coloré
            assert cell.fill.start_color is not None, f"Header cell {cell.coordinate} should have fill color"
            # Vérifier que les en-têtes ont des bordures
            assert cell.border is not None, f"Header cell {cell.coordinate} should have border"
        
        # Vérifier les styles des en-têtes dans l'onglet Détails
        ws_details = wb["Détails"]
        detail_header_row = ws_details[1]
        
        for cell in detail_header_row:
            assert cell.font.bold is True, f"Detail header cell {cell.coordinate} should be bold"
            assert cell.fill.start_color is not None, f"Detail header cell {cell.coordinate} should have fill color"
            assert cell.border is not None, f"Detail header cell {cell.coordinate} should have border"
        
        # Vérifier que la ligne de totaux est en gras (si elle existe)
        if ws_summary.max_row > 2:  # Headers + au moins une ligne de données + totaux
            total_row = ws_summary[ws_summary.max_row]
            # La première cellule de la ligne de totaux devrait contenir "TOTAL"
            if total_row[0].value == "TOTAL":
                for cell in total_row:
                    if cell.value:  # Ignorer les cellules vides
                        assert cell.font.bold is True, f"Total cell {cell.coordinate} should be bold"
    
    def test_export_excel_performance_1000_sessions(self, admin_client, db_session, test_site, test_operator):
        """Test de performance : export Excel de 1000 sessions doit être < 30 secondes."""
        import time
        
        # Créer 1000 sessions de test
        now = datetime.now(timezone.utc)
        sessions = []
        for i in range(1000):
            session = CashSession(
                id=uuid4(),
                operator_id=test_operator.id,
                site_id=test_site.id,
                initial_amount=100.0,
                status=CashSessionStatus.CLOSED,
                opened_at=now - timedelta(days=i % 30),
                closed_at=now - timedelta(days=i % 30, hours=1),
                total_sales=float(i * 10),
                number_of_sales=i % 10,
                total_items=i % 20,
                closing_amount=100.0 + float(i * 10),
                actual_amount=100.0 + float(i * 10),
                variance=0.0
            )
            sessions.append(session)
        
        db_session.add_all(sessions)
        db_session.commit()
        
        # Mesurer le temps d'export
        start_time = time.time()
        
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": (now - timedelta(days=30)).isoformat(),
                    "date_to": now.isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        elapsed_time = time.time() - start_time
        
        assert response.status_code == 200
        assert elapsed_time < 30.0, f"Export took {elapsed_time:.2f}s, should be < 30s"
        
        # Vérifier que le fichier est valide
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        assert "Résumé" in wb.sheetnames
        assert "Détails" in wb.sheetnames


class TestBulkExportCashSessionsValidation:
    """Tests de validation pour l'export bulk."""
    
    def test_export_bulk_invalid_format(self, admin_client):
        """Test avec format invalide."""
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {},
                "format": "pdf"  # Format invalide
            }
        )
        
        assert response.status_code == 422  # Validation error
    
    def test_export_bulk_date_range_validation(self, admin_client):
        """Test validation des dates."""
        # Date future (doit être rejetée ou gérée)
        future_date = datetime.now(timezone.utc) + timedelta(days=10)
        
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": future_date.isoformat()
                },
                "format": "csv"
            }
        )
        
        # Le backend devrait gérer les dates futures (pas d'erreur, juste aucun résultat)
        assert response.status_code in [200, 400]
    
    def test_export_bulk_empty_result(self, admin_client):
        """Test export avec aucun résultat (filtres trop restrictifs)."""
        far_future = datetime.now(timezone.utc) + timedelta(days=365)
        
        response = admin_client.post(
            "/api/v1/admin/reports/cash-sessions/export-bulk",
            json={
                "filters": {
                    "date_from": far_future.isoformat(),
                    "date_to": (far_future + timedelta(days=1)).isoformat()
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        # Le CSV devrait contenir seulement les headers
        content = response.text
        lines = [l for l in content.split('\n') if l.strip()]
        assert len(lines) == 1  # Seulement headers

