"""
Configuration des tests pour l'API Recyclic

Lot C pilote (TEST-01 + extension) : isolation DB pour un sous-ensemble de
modules auth/infra, admin utilisateur, ``test_refresh_token_service`` et
``test_refresh_token_endpoint`` (voir ``_PILOT_DB_ISOLATION_BASENAMES`` et
``_cleanup_pilot_db_tables``). Le
nettoyage s'exécute dans le ``finally`` de ``_db_autouse`` après fermeture
session + connexion du test (évite les verrous SQLite avec un hook teardown).
Un échec de cleanup (inspect schéma ou ``DELETE``) est journalisé et fait
échouer le teardown du test pilote (plus de best-effort silencieux).
"""

import logging
import os
import sys
import tempfile
from pathlib import Path

# Ajouter la racine du projet au PYTHONPATH pour résoudre les imports
# /app/tests/conftest.py -> /app
sys.path.insert(0, str(Path(__file__).parent.parent))

# --- Avant tout import recyclic_api : Settings, engine applicatif, client Redis ---
# Sinon DATABASE_URL / TESTING ne sont pas appliqués par config.py au chargement des modules.
os.environ["TESTING"] = "true"
os.environ.setdefault("SECRET_KEY", "pytest-secret-key-do-not-use-in-production")
os.environ.setdefault("REDIS_URL", "redis://localhost:6379")

_api_dir = Path(__file__).parent.parent
# Fichier SQLite distinct par processus pytest : évite « database is locked » si un autre
# run (ou outil) tient encore pytest_recyclic.db, en particulier sous Windows.
_fd_default_sqlite, _path_default_sqlite = tempfile.mkstemp(
    prefix="pytest_recyclic_", suffix=".db"
)
os.close(_fd_default_sqlite)
_default_sqlite = "sqlite:///" + str(Path(_path_default_sqlite).resolve()).replace("\\", "/")

configured_test_db_url = os.getenv("TEST_DATABASE_URL", "")
db_url = os.getenv("DATABASE_URL", "")

if configured_test_db_url.startswith("postgresql"):
    # Les tests explicitement demandés sur PostgreSQL doivent conserver l'URL fournie.
    test_db_url = configured_test_db_url
elif db_url and not db_url.startswith("sqlite") and "/recyclic" in db_url:
    test_db_url = db_url.rsplit("/", 1)[0] + "/recyclic_test"
else:
    # Sous SQLite, on force toujours un fichier temporaire par processus pytest.
    # Cela évite les collisions de schéma/verrous quand un shell réutilise un ancien
    # TEST_DATABASE_URL ou quand plusieurs suites tournent en parallèle sous Windows.
    test_db_url = _default_sqlite

os.environ["TEST_DATABASE_URL"] = test_db_url
if "DATABASE_URL" not in os.environ:
    os.environ["DATABASE_URL"] = test_db_url

import types

# Précharger reportlab si installé (comme openpyxl ci-dessous) : sinon le stub vide
# casse export_categories PDF (getSampleStyleSheet → {} → KeyError 'Heading1').
try:
    import reportlab  # noqa: F401
except ImportError:
    pass

if "reportlab" not in sys.modules:
    reportlab = types.ModuleType("reportlab")
    lib = types.ModuleType("reportlab.lib")
    colors = types.ModuleType("reportlab.lib.colors")
    colors.HexColor = lambda value: value
    pagesizes = types.ModuleType("reportlab.lib.pagesizes")
    pagesizes.A4 = (0, 0)
    styles = types.ModuleType("reportlab.lib.styles")
    class _Dummy:
        def __init__(self, *args, **kwargs):
            pass

    styles.getSampleStyleSheet = lambda: {}
    styles.ParagraphStyle = _Dummy
    units = types.ModuleType("reportlab.lib.units")
    units.cm = 1
    enums = types.ModuleType("reportlab.lib.enums")
    enums.TA_CENTER = 1
    enums.TA_LEFT = 0
    platypus = types.ModuleType("reportlab.platypus")
    platypus.SimpleDocTemplate = _Dummy
    platypus.Table = _Dummy
    platypus.TableStyle = _Dummy
    platypus.Paragraph = _Dummy
    platypus.Spacer = _Dummy
    platypus.PageBreak = _Dummy
    platypus.KeepTogether = _Dummy
    sys.modules["reportlab"] = reportlab
    sys.modules["reportlab.lib"] = lib
    sys.modules["reportlab.lib.colors"] = colors
    sys.modules["reportlab.lib.pagesizes"] = pagesizes
    sys.modules["reportlab.lib.styles"] = styles
    sys.modules["reportlab.lib.units"] = units
    sys.modules["reportlab.lib.enums"] = enums
    sys.modules["reportlab.platypus"] = platypus

# Stub openpyxl uniquement si le paquet est absent (CI minimale). Si openpyxl est installé,
# le précharger ici évite d'écraser sys.modules avec un Workbook factice (exports Excel en 500).
try:
    import openpyxl  # noqa: F401
except ImportError:
    pass

if "openpyxl" not in sys.modules:
    class _DummyCell:
        def __init__(self):
            self.font = None
            self.fill = None
            self.alignment = None

    class _DummyColumn:
        def __init__(self):
            self.width = None

    class _DummyWorksheet:
        def __init__(self, title: str = ""):
            self.title = title
            self._rows = []
            self.column_dimensions = {chr(ord('A') + i): _DummyColumn() for i in range(26)}

        def append(self, row):
            self._rows.append(row)

        def __getitem__(self, key):
            index = int(key) - 1 if not isinstance(key, int) else key - 1
            row = self._rows[index] if 0 <= index < len(self._rows) else []
            return [_DummyCell() for _ in row]

        @property
        def max_row(self):
            return len(self._rows)

        def iter_rows(self, min_row=1, max_row=None):
            max_row = max_row or self.max_row
            for idx in range(min_row - 1, max_row):
                row = self._rows[idx] if 0 <= idx < len(self._rows) else []
                yield [_DummyCell() for _ in row]

    class _DummyWorkbook:
        """Minimal Workbook stub : ``sheetnames``, ``remove``, ``create_sheet`` (exports bulk Excel)."""

        def __init__(self):
            self._sheets: dict[str, _DummyWorksheet] = {"Sheet": _DummyWorksheet("Sheet")}
            self.active = self._sheets["Sheet"]

        @property
        def sheetnames(self) -> list[str]:
            return list(self._sheets.keys())

        def __getitem__(self, name: str) -> _DummyWorksheet:
            return self._sheets[name]

        def remove(self, worksheet: _DummyWorksheet) -> None:
            for key, ws in list(self._sheets.items()):
                if ws is worksheet:
                    del self._sheets[key]
                    return

        def create_sheet(self, title: str) -> _DummyWorksheet:
            sh = _DummyWorksheet(title)
            self._sheets[title] = sh
            return sh

        def save(self, _buffer):
            pass

    class _DummyFont:
        def __init__(self, *args, **kwargs):
            pass

    class _DummyAlignment:
        def __init__(self, *args, **kwargs):
            pass

    class _DummyPatternFill:
        def __init__(self, *args, **kwargs):
            pass

    class _DummyBorder:
        def __init__(self, *args, **kwargs):
            pass

    class _DummySide:
        def __init__(self, *args, **kwargs):
            pass

    def _dummy_load_workbook(*args, **kwargs):
        """Mock pour load_workbook qui retourne un workbook vide."""
        return _DummyWorkbook()

    openpyxl = types.ModuleType("openpyxl")
    styles_module = types.ModuleType("openpyxl.styles")
    styles_module.Font = _DummyFont
    styles_module.Alignment = _DummyAlignment
    styles_module.PatternFill = _DummyPatternFill
    styles_module.Border = _DummyBorder
    styles_module.Side = _DummySide

    sys.modules["openpyxl"] = openpyxl
    sys.modules["openpyxl.styles"] = styles_module
    openpyxl.styles = styles_module

    def _workbook_factory():
        return _DummyWorkbook()

    openpyxl.Workbook = _DummyWorkbook
    openpyxl.load_workbook = _dummy_load_workbook

import re

import pytest
import pytest_asyncio

# Module `clean_legacy_import` absent (script hors dépôt) — évite l'échec de collecte.
collect_ignore = ["test_clean_legacy_import.py"]

from fastapi.testclient import TestClient
import httpx
from httpx import AsyncClient
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.orm import sessionmaker
import json
from typing import Generator
import uuid

from recyclic_api.main import app
from recyclic_api.core.database import get_db, Base
from sqlalchemy.orm import Session

from recyclic_api.models.user import User, UserRole, UserStatus
from tests.memory_redis_for_tests import MemoryRedisForTests
from recyclic_api.models.login_history import LoginHistory
from recyclic_api.models.site import Site
from recyclic_api.models.cash_register import CashRegister
from recyclic_api.models.deposit import Deposit
from recyclic_api.models.payment_method import PaymentMethodDefinition
from recyclic_api.models.accounting_config import (
    GLOBAL_ACCOUNTING_SETTINGS_ROW_ID,
    AccountingConfigRevision,
    GlobalAccountingSettings,
)
from recyclic_api.models.accounting_period_authority import AccountingPeriodAuthoritySnapshot
from recyclic_api.models.sale import Sale
from recyclic_api.models.sale_reversal import SaleReversal
from recyclic_api.models.exceptional_refund import ExceptionalRefund
from recyclic_api.models.cash_disbursement import CashDisbursement
from recyclic_api.models.material_exchange import MaterialExchange
from recyclic_api.models.sale_item import SaleItem
from recyclic_api.models.cash_session import CashSession
from recyclic_api.models.paheko_outbox import PahekoOutboxItem
from recyclic_api.models.paheko_outbox_sync_transition import PahekoOutboxSyncTransition
from recyclic_api.models.paheko_cash_session_close_mapping import PahekoCashSessionCloseMapping
from recyclic_api.models.payment_transaction import PaymentTransaction
from recyclic_api.models.sync_log import SyncLog
from recyclic_api.models.registration_request import RegistrationRequest
from recyclic_api.models.user_status_history import UserStatusHistory
from recyclic_api.models.admin_setting import AdminSetting  # noqa: F401 — admin_settings (SQLite)
from recyclic_api.models.category import Category
from recyclic_api.models.legacy_category_mapping_cache import LegacyCategoryMappingCache
from recyclic_api.models.preset_button import PresetButton  # noqa: F401 — preset_buttons (tests presets API)
from recyclic_api.models.poste_reception import PosteReception
from recyclic_api.models.ticket_depot import TicketDepot
from recyclic_api.models.ligne_depot import LigneDepot
from recyclic_api.models.user_session import UserSession  # noqa: F401 — table login / refresh
from recyclic_api.models.setting import Setting  # noqa: F401 — refresh_token_max_hours (RefreshTokenService)
from recyclic_api.models.email_log import EmailLog  # noqa: F401 — tests email_logs (SQLite partiel)
from recyclic_api.models.permission import (  # noqa: F401 — tests groupes / permissions (SQLite)
    Group,
    Permission,
    group_permissions,
    user_groups,
)
from recyclic_api.core.security import create_access_token, hash_password

logger = logging.getLogger(__name__)

# httpx >= 0.28 : ``AsyncClient(app=...)`` supprimé — transport ASGI explicite.
def _async_client_for_fastapi_app():
    """Construit un ``AsyncClient`` contre l'app FastAPI, compatible httpx 0.25–0.28+."""
    base_url = "http://testserver"
    try:
        return AsyncClient(app=app, base_url=base_url)
    except TypeError:
        # httpx 0.28+ : uniquement ``transport=ASGITransport(app=...)``.
        return AsyncClient(
            transport=httpx.ASGITransport(app=app),
            base_url=base_url,
        )

# SQLite ne conserve pas le tz des colonnes DateTime(timezone=True) : l'ORM peut
# renvoyer des datetimes naïfs, ce qui casse les comparaisons avec ``datetime.now(timezone.utc)``.
_USER_SESSION_UTC_ATTRS = (
    "expires_at",
    "issued_at",
    "last_used_at",
    "revoked_at",
    "created_at",
    "updated_at",
)


def _coerce_user_session_datetimes_utc(target: UserSession) -> None:
    from datetime import timezone as _tz

    for attr in _USER_SESSION_UTC_ATTRS:
        val = getattr(target, attr, None)
        if val is not None and getattr(val, "tzinfo", None) is None:
            setattr(target, attr, val.replace(tzinfo=_tz.utc))

# Modules du pilote isolation DB (Lot C / TEST-01 + TEST-02 admin) : nettoyage post-test ciblé.
_PILOT_DB_ISOLATION_BASENAMES = frozenset(
    {
        "test_infrastructure.py",
        "test_auth_login_endpoint.py",
        "test_auth_logging.py",
        "test_auth_inactive_user_middleware.py",
        "test_auth_login_username_password.py",
        "test_admin_user_status_endpoint.py",
        "test_admin_user_management.py",
        "test_refresh_token_service.py",
        "test_refresh_token_endpoint.py",
    }
)

# Ordre : enfants (FK vers users) puis users. Ignoré si la table n'existe pas (SQLite minimal).
# ``settings`` : clés globales (ex. refresh_token_max_hours) — évite qu'un test pilote
# impose la valeur au suivant après commit + DELETE users/sessions.
_PILOT_CLEANUP_TABLES_ORDER = (
    "audit_logs",
    "user_sessions",
    "login_history",
    "user_status_history",
    "settings",
    "users",
)

SQLALCHEMY_DATABASE_URL = os.environ["TEST_DATABASE_URL"]

engine_kwargs = {}
if SQLALCHEMY_DATABASE_URL.startswith("sqlite"):
    # timeout : file-lock SQLite quand le pilote enchaîne connexion test + nettoyage
    engine_kwargs["connect_args"] = {"check_same_thread": False, "timeout": 30}

engine = create_engine(SQLALCHEMY_DATABASE_URL, **engine_kwargs)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Même Engine / SessionLocal que ``recyclic_api.core.database`` : évite une base « vide »
# pour ``get_token_expiration_minutes`` (``next(get_db())`` hors override FastAPI) et
# aligne ``check_same_thread=False`` pour TestClient / httpx (threads).
import recyclic_api.core.database as _app_database_module

try:
    _app_database_module.engine.dispose(close=True)
except Exception:
    pass
_app_database_module.engine = engine
_app_database_module.SessionLocal = TestingSessionLocal

if SQLALCHEMY_DATABASE_URL.startswith("sqlite"):
    from sqlalchemy import event

    @event.listens_for(UserSession, "load")
    def _sqlite_user_session_load_utc(target, context):  # noqa: ARG001
        _coerce_user_session_datetimes_utc(target)

    @event.listens_for(UserSession, "refresh")
    def _sqlite_user_session_refresh_utc(target, context, attrs):  # noqa: ARG001
        _coerce_user_session_datetimes_utc(target)


def _cleanup_pilot_db_tables() -> None:
    """
    Supprime les lignes des tables les plus touchées par auth/login dans le pilote.

    Objectif : annuler les effets des ``commit()`` applicatifs entre tests du pilote,
    sans refactorer toute la suite. Limite : exécution mixte avec d'autres tests
    dans le même processus pytest peut voir ces tables vidées après un test pilote
    (voir doc équipe / résumé Lot C).

    Chaque ``DELETE`` s'exécute dans sa propre transaction : un échec n'empêche pas
    de tenter les tables suivantes, et n'abort pas silencieusement une transaction
    globale (comportement critique sous PostgreSQL). Tout échec est loggé ; au
    moins une erreur lève ``RuntimeError`` pour que le teardown pilote soit visible.
    """
    from sqlalchemy import inspect

    try:
        insp = inspect(engine)
        existing = set(insp.get_table_names())
    except Exception as e:
        logger.error(
            "Pilote Lot C (cleanup): impossible d'inspecter le schéma: %s",
            e,
            exc_info=True,
        )
        raise RuntimeError(
            "Pilote Lot C (cleanup): échec inspect(engine) — isolation non garantie."
        ) from e

    delete_errors: list[tuple[str, Exception]] = []
    for table in _PILOT_CLEANUP_TABLES_ORDER:
        if table not in existing:
            continue
        try:
            with engine.begin() as conn:
                conn.execute(text(f"DELETE FROM {table}"))
        except Exception as e:
            logger.error(
                "Pilote Lot C (cleanup): DELETE FROM %s a échoué: %s",
                table,
                e,
                exc_info=True,
            )
            delete_errors.append((table, e))

    if delete_errors:
        summary = "; ".join(f"{name}: {err}" for name, err in delete_errors)
        raise RuntimeError(
            f"Pilote Lot C (cleanup): {len(delete_errors)} DELETE en échec — {summary}"
        ) from delete_errors[0][1]

    # Cache classe ActivityService (seuil lu depuis ``settings``) : sans reset, une valeur
    # mise en cache peut survivre au DELETE ``settings`` jusqu'à expiration TTL (60 s).
    try:
        from recyclic_api.services.activity_service import (
            ActivityService,
            DEFAULT_ACTIVITY_THRESHOLD_MINUTES,
        )

        ActivityService._cached_threshold_minutes = DEFAULT_ACTIVITY_THRESHOLD_MINUTES
        ActivityService._cache_expiration_timestamp = 0.0
    except Exception as e:
        logger.error(
            "Pilote Lot C (cleanup): impossible de réinitialiser le cache ActivityService: %s",
            e,
            exc_info=True,
        )
        raise RuntimeError(
            "Pilote Lot C (cleanup): reset du cache ActivityService en échec — isolation non garantie."
        ) from e


def _request_is_pilot_db_isolation(request) -> bool:
    try:
        path = getattr(request.node, "path", None) or getattr(request.node, "fspath", None)
        name = path.name if path is not None else ""
    except Exception:
        return False
    return name in _PILOT_DB_ISOLATION_BASENAMES


def _sqlite_align_sales_story_63(bind) -> None:
    """SQLite partiel : ajoute ``sales.lifecycle_status`` si la table existait sans colonne (Story 6.3)."""
    if not str(bind.url).startswith("sqlite"):
        return
    with bind.connect() as conn:
        insp = inspect(conn)
        if not insp.has_table("sales"):
            return
        cols = {c["name"] for c in insp.get_columns("sales")}
        if "lifecycle_status" not in cols:
            conn.execute(
                text(
                    "ALTER TABLE sales ADD COLUMN lifecycle_status VARCHAR(32) NOT NULL DEFAULT 'completed'"
                )
            )
            conn.commit()


def _sqlite_align_sales_story_65(bind) -> None:
    """SQLite partiel : colonnes Story 6.5 (encaissements sans article) si table sales préexistante."""
    if not str(bind.url).startswith("sqlite"):
        return
    with bind.connect() as conn:
        insp = inspect(conn)
        if not insp.has_table("sales"):
            return
        cols = {c["name"] for c in insp.get_columns("sales")}
        if "special_encaissement_kind" not in cols:
            conn.execute(text("ALTER TABLE sales ADD COLUMN special_encaissement_kind VARCHAR(64)"))
        if "adherent_reference" not in cols:
            conn.execute(text("ALTER TABLE sales ADD COLUMN adherent_reference VARCHAR(200)"))
        if "social_action_kind" not in cols:
            conn.execute(text("ALTER TABLE sales ADD COLUMN social_action_kind VARCHAR(64)"))
        conn.commit()


def _sqlite_align_groups_story_23(bind) -> None:
    """SQLite : `create_all` ne rajoute pas les colonnes sur une table existante (Story 2.3)."""
    if not str(bind.url).startswith("sqlite"):
        return
    with bind.connect() as conn:
        insp = inspect(conn)
        if not insp.has_table("groups"):
            return
        cols = {c["name"] for c in insp.get_columns("groups")}
        if "key" not in cols:
            conn.execute(text("ALTER TABLE groups ADD COLUMN key VARCHAR"))
        if "display_name" not in cols:
            conn.execute(text("ALTER TABLE groups ADD COLUMN display_name VARCHAR"))
        if "site_id" not in cols:
            conn.execute(text("ALTER TABLE groups ADD COLUMN site_id VARCHAR"))
        conn.commit()

        def _slug(label: str) -> str:
            s = re.sub(r"[^a-z0-9]+", "-", (label or "").lower().strip())
            s = re.sub(r"-+", "-", s).strip("-")
            return s or "group"

        rows = conn.execute(
            text("SELECT id, name FROM groups WHERE key IS NULL OR key = ''")
        ).fetchall()
        used = {
            r[0]
            for r in conn.execute(
                text("SELECT key FROM groups WHERE key IS NOT NULL AND key != ''")
            ).fetchall()
        }
        for rid, name in rows:
            base = _slug(str(name) if name else "")
            cand = base
            n = 0
            while cand in used:
                n += 1
                cand = f"{base}-{n}"
            used.add(cand)
            conn.execute(
                text(
                    "UPDATE groups SET key = :k, display_name = COALESCE(display_name, :dn) WHERE id = :id"
                ),
                {"k": cand, "dn": str(name) if name else "", "id": rid},
            )
        conn.commit()


def _sqlite_align_paheko_outbox_story_82(bind) -> None:
    """SQLite : colonnes Story 8.2 (`next_retry_at`, `rejection_reason`) si table outbox préexistante."""
    if not str(bind.url).startswith("sqlite"):
        return
    with bind.connect() as conn:
        insp = inspect(conn)
        if not insp.has_table("paheko_outbox_items"):
            return
        cols = {c["name"] for c in insp.get_columns("paheko_outbox_items")}
        if "next_retry_at" not in cols:
            conn.execute(text("ALTER TABLE paheko_outbox_items ADD COLUMN next_retry_at DATETIME"))
            conn.execute(
                text(
                    "CREATE INDEX IF NOT EXISTS ix_paheko_outbox_next_retry_at "
                    "ON paheko_outbox_items (next_retry_at)"
                )
            )
        if "rejection_reason" not in cols:
            conn.execute(text("ALTER TABLE paheko_outbox_items ADD COLUMN rejection_reason TEXT"))
        if "mapping_resolution_error" not in cols:
            conn.execute(text("ALTER TABLE paheko_outbox_items ADD COLUMN mapping_resolution_error VARCHAR(64)"))
            conn.execute(
                text(
                    "CREATE INDEX IF NOT EXISTS ix_paheko_outbox_mapping_resolution_error "
                    "ON paheko_outbox_items (mapping_resolution_error)"
                )
            )
        conn.commit()


def _sqlite_align_accounting_period_authority_story_225(bind) -> None:
    """SQLite : table + seed autorité exercice + permission second parcours (Story 22.5)."""
    if not str(bind.url).startswith("sqlite"):
        return

    from datetime import datetime, timezone

    from sqlalchemy import inspect, text

    AccountingPeriodAuthoritySnapshot.__table__.create(bind, checkfirst=True)

    with bind.connect() as conn:
        insp = inspect(conn)
        if not insp.has_table("accounting_period_authority_snapshots"):
            conn.commit()
            return
        row_ct = conn.execute(text("SELECT COUNT(*) FROM accounting_period_authority_snapshots")).scalar() or 0
        if row_ct == 0:
            y = datetime.now(timezone.utc).year
            conn.execute(
                text(
                    "INSERT INTO accounting_period_authority_snapshots "
                    "(id, current_open_fiscal_year, fetched_at, source, version) "
                    "VALUES (:id, :y, :ts, 'local_test_seed', 1)"
                ),
                {
                    "id": "00000000-0000-5000-8000-000000000001",
                    "y": y,
                    "ts": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
                },
            )
        perm_ct = conn.execute(
            text("SELECT COUNT(*) FROM permissions WHERE name = 'accounting.prior_year_refund'")
        ).scalar() or 0
        if perm_ct == 0:
            conn.execute(
                text(
                    "INSERT INTO permissions (id, name, description) VALUES "
                    "(:pid, 'accounting.prior_year_refund', 'Story 22.5 second parcours N-1 clos')"
                ),
                {"pid": str(uuid.uuid4())},
            )
        conn.commit()


def _sqlite_align_accounting_expert_story_223(bind) -> None:
    """SQLite : tables révision 22.3, seed minimal, FK session (alignement tests sans Alembic)."""
    if not str(bind.url).startswith("sqlite"):
        return

    GlobalAccountingSettings.__table__.create(bind, checkfirst=True)
    AccountingConfigRevision.__table__.create(bind, checkfirst=True)

    _gid = GLOBAL_ACCOUNTING_SETTINGS_ROW_ID
    snapshot_to_seed = None
    with bind.connect() as conn:
        insp = inspect(conn)
        if insp.has_table("global_accounting_settings"):
            c = conn.execute(text("SELECT COUNT(*) FROM global_accounting_settings WHERE id = :id"), {"id": str(_gid)}).scalar()
            if c == 0:
                conn.execute(
                    text(
                        "INSERT INTO global_accounting_settings (id, default_sales_account, "
                        "default_donation_account, prior_year_refund_account, cash_journal_code, default_entry_label_prefix) "
                        "VALUES (:id, '707', '7541', '672', NULL, 'Z caisse')"
                    ),
                    {"id": str(_gid)},
                )
        rev_ct = 0
        if insp.has_table("accounting_config_revisions"):
            rev_ct = conn.execute(text("SELECT COUNT(*) FROM accounting_config_revisions")).scalar() or 0
        if rev_ct == 0:
            pm_snapshot = []
            if insp.has_table("payment_methods"):
                rows = conn.execute(
                    text(
                        "SELECT id, code, label, active, kind, paheko_debit_account, paheko_refund_credit_account, "
                        "min_amount, max_amount, display_order, notes, archived_at FROM payment_methods "
                        "ORDER BY display_order, code"
                    )
                ).fetchall()
                for r in rows:
                    archived = r[11].isoformat() if r[11] is not None else None
                    pm_snapshot.append(
                        {
                            "id": str(r[0]),
                            "code": r[1],
                            "label": r[2],
                            "active": bool(r[3]),
                            "kind": r[4],
                            "paheko_debit_account": r[5],
                            "paheko_refund_credit_account": r[6],
                            "min_amount": r[7],
                            "max_amount": r[8],
                            "display_order": r[9],
                            "notes": r[10],
                            "archived_at": archived,
                        }
                    )
            gcols = {c["name"] for c in insp.get_columns("global_accounting_settings")}
            if {"cash_journal_code", "default_entry_label_prefix"}.issubset(gcols):
                g = conn.execute(
                    text(
                        "SELECT default_sales_account, default_donation_account, prior_year_refund_account, "
                        "cash_journal_code, default_entry_label_prefix "
                        "FROM global_accounting_settings WHERE id = :id"
                    ),
                    {"id": str(_gid)},
                ).fetchone()
            else:
                g = conn.execute(
                    text(
                        "SELECT default_sales_account, default_donation_account, prior_year_refund_account "
                        "FROM global_accounting_settings WHERE id = :id"
                    ),
                    {"id": str(_gid)},
                ).fetchone()
            if g is None:
                ga = ("707", "7541", "672", "", "Z caisse")
            elif len(g) >= 5:
                ga = (g[0], g[1], g[2], (g[3] or "").strip() if g[3] is not None else "", str(g[4] or "Z caisse"))
            else:
                ga = (g[0], g[1], g[2], "", "Z caisse")
            snapshot_to_seed = {
                "schema_version": 1,
                "global_accounts": {
                    "default_sales_account": ga[0],
                    "default_donation_account": ga[1],
                    "prior_year_refund_account": ga[2],
                    "cash_journal_code": ga[3],
                    "default_entry_label_prefix": ga[4],
                },
                "payment_methods": pm_snapshot,
            }
        if insp.has_table("cash_sessions"):
            cols = {c["name"] for c in insp.get_columns("cash_sessions")}
            if "accounting_config_revision_id" not in cols:
                conn.execute(
                    text("ALTER TABLE cash_sessions ADD COLUMN accounting_config_revision_id VARCHAR(36)")
                )
            # Story 22.6 — colonne JSON snapshot (SQLite tests hors alembic complet)
            if "accounting_close_snapshot" not in cols:
                conn.execute(text("ALTER TABLE cash_sessions ADD COLUMN accounting_close_snapshot JSON"))
        conn.commit()

    if snapshot_to_seed is not None:
        from sqlalchemy.orm import sessionmaker

        ORMSession = sessionmaker(bind=bind)
        seed_sess = ORMSession()
        try:
            rev = AccountingConfigRevision(
                id=uuid.uuid4(),
                revision_seq=1,
                snapshot_json=json.dumps(snapshot_to_seed),
                note="SQLite test seed Story 22.3",
            )
            seed_sess.add(rev)
            seed_sess.commit()
        finally:
            seed_sess.close()


def _sqlite_align_payment_canonical_story_221(bind) -> None:
    """SQLite : ajoute le socle canonique `payment_methods` / `payment_transactions` si le schéma existait déjà."""
    if not str(bind.url).startswith("sqlite"):
        return

    PaymentMethodDefinition.__table__.create(bind, checkfirst=True)

    with bind.connect() as conn:
        insp = inspect(conn)
        if not insp.has_table("payment_transactions"):
            return
        cols = {c["name"] for c in insp.get_columns("payment_transactions")}
        if "payment_method_id" not in cols:
            conn.execute(text("ALTER TABLE payment_transactions ADD COLUMN payment_method_id VARCHAR(36)"))
        if "nature" not in cols:
            conn.execute(text("ALTER TABLE payment_transactions ADD COLUMN nature VARCHAR(32)"))
        if "direction" not in cols:
            conn.execute(text("ALTER TABLE payment_transactions ADD COLUMN direction VARCHAR(32)"))
        if "original_sale_id" not in cols:
            conn.execute(text("ALTER TABLE payment_transactions ADD COLUMN original_sale_id VARCHAR(36)"))
        if "original_payment_method_id" not in cols:
            conn.execute(text("ALTER TABLE payment_transactions ADD COLUMN original_payment_method_id VARCHAR(36)"))
        if "is_prior_year_special_case" not in cols:
            conn.execute(
                text(
                    "ALTER TABLE payment_transactions "
                    "ADD COLUMN is_prior_year_special_case BOOLEAN NOT NULL DEFAULT 0"
                )
            )
        if "paheko_account_override" not in cols:
            conn.execute(text("ALTER TABLE payment_transactions ADD COLUMN paheko_account_override VARCHAR(32)"))
        if "notes" not in cols:
            conn.execute(text("ALTER TABLE payment_transactions ADD COLUMN notes TEXT"))
        conn.execute(
            text(
                "CREATE INDEX IF NOT EXISTS ix_payment_transactions_payment_method_id "
                "ON payment_transactions (payment_method_id)"
            )
        )
        conn.execute(
            text(
                "CREATE INDEX IF NOT EXISTS ix_payment_transactions_original_sale_id "
                "ON payment_transactions (original_sale_id)"
            )
        )
        conn.commit()


# Fonction pour créer les tables
def create_tables_if_not_exist():
    """Créer les tables si elles n'existent pas"""
    try:
        print("[tests] Creation des tables...")
        # Créer les types ENUM PostgreSQL nécessaires avant de créer les tables
        # (Base.metadata.create_all() ne les crée pas automatiquement)
        if SQLALCHEMY_DATABASE_URL.startswith("postgresql"):
            with engine.connect() as conn:
                # Créer le type ENUM cashsessionstep s'il n'existe pas
                conn.execute(text("""
                    DO $$ BEGIN
                        CREATE TYPE cashsessionstep AS ENUM ('ENTRY', 'SALE', 'EXIT');
                    EXCEPTION
                        WHEN duplicate_object THEN null;
                    END $$;
                """))
                conn.commit()
        if SQLALCHEMY_DATABASE_URL.startswith("sqlite"):
            # Schéma SQLite partiel : tables listées ci-dessous (JSONB non supporté par le dialecte SQLite,
            # ex. audit_logs — voir fixture _sqlite_skip_audit_log_commit).
            Base.metadata.create_all(
                bind=engine,
                tables=[
                    User.__table__,
                    UserStatusHistory.__table__,
                    LoginHistory.__table__,
                    UserSession.__table__,
                    Setting.__table__,
                    EmailLog.__table__,
                    Permission.__table__,
                    Group.__table__,
                    group_permissions,
                    user_groups,
                    # Sessions caisse / ventes (tests admin maintenance, cash_session_*, etc.)
                    Site.__table__,
                    # FK site_id — requis pour DELETE sites (UoW SQLAlchemy charge la relation)
                    RegistrationRequest.__table__,
                    AdminSetting.__table__,
                    CashRegister.__table__,
                    GlobalAccountingSettings.__table__,
                    AccountingConfigRevision.__table__,
                    AccountingPeriodAuthoritySnapshot.__table__,
                    CashSession.__table__,
                    PahekoOutboxItem.__table__,
                    PahekoOutboxSyncTransition.__table__,
                    PahekoCashSessionCloseMapping.__table__,
                    PaymentMethodDefinition.__table__,
                    Sale.__table__,
                    SaleReversal.__table__,
                    ExceptionalRefund.__table__,
                    MaterialExchange.__table__,
                    CashDisbursement.__table__,
                    SaleItem.__table__,
                    PaymentTransaction.__table__,
                    Deposit.__table__,
                    # Réception / dépôt (test_reception_live_stats, stats live)
                    Category.__table__,
                    LegacyCategoryMappingCache.__table__,
                    PresetButton.__table__,
                    PosteReception.__table__,
                    TicketDepot.__table__,
                    LigneDepot.__table__,
                ],
            )
            _sqlite_align_sales_story_63(engine)
            _sqlite_align_sales_story_65(engine)
            _sqlite_align_groups_story_23(engine)
            _sqlite_align_paheko_outbox_story_82(engine)
            _sqlite_align_payment_canonical_story_221(engine)
            _sqlite_align_accounting_expert_story_223(engine)
            _sqlite_align_accounting_period_authority_story_225(engine)
        else:
            Base.metadata.create_all(bind=engine)
        print("[tests] Tables creees avec succes")
    except Exception as e:
        print(f"[tests] Erreur lors de la creation des tables: {e}")

@pytest.fixture(scope="session")
def db_engine():
    """Crée les tables une seule fois pour toute la session de test."""
    print("[tests] Moteur de base de donnees de test pret.")
    create_tables_if_not_exist()
    return engine


@pytest.fixture(autouse=True)
def _sqlite_skip_audit_log_commit(request):
    """
    Sous SQLite, audit_logs (JSONB) n'est pas créé ; log_audit() et dérivés
    (log_role_change, etc.) font un commit() ou échouent sur insert. On neutralise
    log_audit dans core.audit (imports dynamiques ex. users), auth, admin_users_credentials,
    admin_activity_threshold, reception (PATCH poids ligne) et sale_service
    (rollback d'insert audit annulerait la session SQLite).
    """
    if not SQLALCHEMY_DATABASE_URL.startswith("sqlite"):
        yield
        return
    if request.node.get_closest_marker("no_db"):
        yield
        return
    from unittest.mock import patch

    from recyclic_api.core import audit as audit_core
    from recyclic_api.api.api_v1.endpoints import (
        admin_users_credentials as admin_users_credentials_endpoints,
    )
    from recyclic_api.api.api_v1.endpoints import admin_activity_threshold as admin_activity_threshold_endpoints
    from recyclic_api.api.api_v1.endpoints import auth as auth_endpoints
    from recyclic_api.api.api_v1.endpoints import reception as reception_endpoints
    from recyclic_api.services import cash_disbursement_service as cash_disbursement_service_mod
    from recyclic_api.services import exceptional_refund_service as exceptional_refund_service_mod
    from recyclic_api.services import material_exchange_service as material_exchange_service_mod
    from recyclic_api.services import sale_service as sale_service_mod

    _noop = lambda *args, **kwargs: None

    with (
        patch.object(audit_core, "log_audit", _noop),
        patch.object(auth_endpoints, "log_audit", _noop),
        patch.object(admin_users_credentials_endpoints, "log_audit", _noop),
        patch.object(admin_activity_threshold_endpoints, "log_audit", _noop),
        patch.object(reception_endpoints, "log_audit", _noop),
        patch.object(sale_service_mod, "log_audit", _noop),
        patch.object(exceptional_refund_service_mod, "log_audit", _noop),
        patch.object(material_exchange_service_mod, "log_audit", _noop),
        patch.object(cash_disbursement_service_mod, "log_audit", _noop),
    ):
        yield

@pytest.fixture(scope="function", autouse=True)
def _db_autouse(db_engine, request):
    """
    Prépare, pour chaque test, une connexion SQLAlchemy + une Session ORM et branche l'app dessus.

    **Ce que cette fixture garantit**
    - Une connexion dédiée au test et une ``Session`` liée à cette connexion.
    - Au début du test, une transaction est ouverte sur la connexion via ``connection.begin()`` ;
      le retour n'est pas utilisé dans le teardown (**ni commit ni rollback explicites** sur cet
      objet — comportement historique volontairement inchangé).
    - L'override FastAPI ``get_db`` ouvre une **nouvelle** ``Session`` par requête HTTP
      (même ``connection``) ; la ``Session`` ``yield``-ée au test reste celle du code de test.

    **Ce qu'elle ne garantit pas**
    - **Pas d'isolation « rollback entre tests »** via cette fixture seule : tout ``commit()``
      laisse les données visibles sur la connexion / le moteur tant que la connexion vit.
    - **Pilote Lot C (TEST-01)** : pour les modules listés dans
      ``_PILOT_DB_ISOLATION_BASENAMES``, après fermeture de la connexion du test, un
      ``DELETE`` ciblé sur ``audit_logs``, ``user_sessions``, ``login_history``,
      ``user_status_history``, ``settings``, ``users``
      (tables présentes uniquement). Cela compense les commits du login / audit pour ce
      sous-ensemble seulement ; ce n'est pas un rollback transactionnel et d'autres tables
      peuvent encore conserver des lignes si un test pilote les remplit.

    Tests marqués ``@pytest.mark.no_db`` : pas de setup DB, la fixture produit ``None``.
    """
    if request.node.get_closest_marker("no_db"):
        yield None
        return

    connection = db_engine.connect()
    # Transaction racine dédiée au test : on la rollback explicitement au teardown
    # pour éviter les fuites d'état et relâcher les verrous SQLite entre tests.
    transaction = connection.begin()

    session = TestingSessionLocal(bind=connection)

    def override_get_db():
        # Une Session par requête HTTP : le TestClient Starlette exécute les routes
        # synchrones dans un thread pool ; réutiliser la même Session ORM entre threads
        # est indéfini (lectures vides après commits, export ticket CSV, etc.).
        req_session = TestingSessionLocal(bind=connection)
        try:
            yield req_session
        finally:
            req_session.close()

    # Override ciblé (pas de clear global)
    app.dependency_overrides[get_db] = override_get_db
    try:
        yield session
    finally:
        app.dependency_overrides.pop(get_db, None)
        session.close()
        if transaction.is_active:
            transaction.rollback()
        connection.close()
        # Pilote TEST-01 : nettoyage après libération de la connexion du test (ordre critique SQLite).
        if _request_is_pilot_db_isolation(request):
            _cleanup_pilot_db_tables()

@pytest.fixture(scope="function")
def db_session(_db_autouse):
    """Même session que ``_db_autouse`` (alias explicite pour les tests). Voir docstring de ``_db_autouse``."""
    return _db_autouse

@pytest.fixture(scope="function")
def client(db_session):
    """Fixture pour le client de test FastAPI qui utilise la session de test."""
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture(scope="function")
def client_with_jwt_auth(db_session):
    """Fixture pour un client de test avec utilisateur authentifié (compatible avec les tests existants)."""
    with TestClient(app) as c:
        yield c


@pytest.fixture(scope="function")
def admin_client(db_session: Session) -> Generator[TestClient, None, None]:
    """
    Fixture pour un client de test avec les droits d'administrateur.
    Crée un utilisateur admin, génère un token JWT et configure le client.
    """
    admin_username = f"admin_{uuid.uuid4().hex}@test.com"
    admin_password = "admin_password"
    hashed_password = hash_password(admin_password)

    admin_user = User(
        username=admin_username,
        hashed_password=hashed_password,
        hashed_pin=hash_password("1234"),
        role=UserRole.ADMIN,
        status=UserStatus.ACTIVE,
        legacy_external_contact_id=f"admin-fixture-{uuid.uuid4().hex[:12]}",
    )
    db_session.add(admin_user)
    db_session.commit()
    db_session.refresh(admin_user)

    # Génération du token
    access_token = create_access_token(data={"sub": str(admin_user.id)})

    # Configuration du client de test avec headers par défaut
    # TestClient n'utilise pas automatiquement client.headers, il faut passer headers dans chaque requête
    # On crée un wrapper qui injecte automatiquement le header
    class AuthenticatedTestClient:
        """Wrapper pour TestClient qui injecte automatiquement le header Authorization."""
        def __init__(self, client: TestClient, token: str):
            self._client = client
            self._auth_header = {"Authorization": f"Bearer {token}"}
        
        def get(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.get(url, headers=headers, **kwargs)
        
        def post(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.post(url, headers=headers, **kwargs)
        
        def patch(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.patch(url, headers=headers, **kwargs)
        
        def put(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.put(url, headers=headers, **kwargs)
        
        def delete(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.delete(url, headers=headers, **kwargs)
    
    with TestClient(app) as client:
        authenticated_client = AuthenticatedTestClient(client, access_token)
        yield authenticated_client


@pytest.fixture(scope="function")
def user_client(db_session: Session) -> Generator[TestClient, None, None]:
    """Client authentifié avec rôle USER (permissions sites : liste OK, CRUD admin refusé)."""
    user_username = f"user_{uuid.uuid4().hex}@test.com"
    user_password = "user_password"
    hashed_password = hash_password(user_password)

    user = User(
        username=user_username,
        hashed_password=hashed_password,
        role=UserRole.USER,
        status=UserStatus.ACTIVE,
        legacy_external_contact_id="777777777",
    )
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)

    access_token = create_access_token(data={"sub": str(user.id)})

    class AuthenticatedTestClient:
        """Wrapper pour TestClient qui injecte automatiquement le header Authorization."""

        def __init__(self, client: TestClient, token: str):
            self._client = client
            self._auth_header = {"Authorization": f"Bearer {token}"}

        def get(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.get(url, headers=headers, **kwargs)

        def post(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.post(url, headers=headers, **kwargs)

        def patch(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.patch(url, headers=headers, **kwargs)

        def put(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.put(url, headers=headers, **kwargs)

        def delete(self, url: str, **kwargs):
            headers = kwargs.pop("headers", {})
            headers.update(self._auth_header)
            return self._client.delete(url, headers=headers, **kwargs)

    with TestClient(app) as client:
        authenticated_client = AuthenticatedTestClient(client, access_token)
        yield authenticated_client


@pytest.fixture(scope="function")
def super_admin_client(db_session: Session) -> Generator[TestClient, None, None]:
    """
    Fixture pour un client de test avec les droits de super-administrateur.
    Crée un utilisateur super-admin, génère un token JWT et configure le client.
    """
    # Création de l'utilisateur super-admin
    super_admin_username = f"superadmin_{uuid.uuid4().hex}@test.com"
    super_admin_password = "superadmin_password"
    hashed_password = hash_password(super_admin_password)

    super_admin_user = User(
        username=super_admin_username,
        hashed_password=hashed_password,
        hashed_pin=hash_password("1234"),
        role=UserRole.SUPER_ADMIN,
        status=UserStatus.ACTIVE,
        legacy_external_contact_id="888888888",
    )
    db_session.add(super_admin_user)
    db_session.commit()
    db_session.refresh(super_admin_user)

    # Génération du token
    access_token = create_access_token(data={"sub": str(super_admin_user.id)})

    # Configuration du client de test
    with TestClient(app) as client:
        client.headers["Authorization"] = f"Bearer {access_token}"
        yield client

@pytest_asyncio.fixture(scope="function")
async def async_client():
    """Fixture pour le client de test asynchrone FastAPI (httpx 0.25 ``app=`` ou 0.28+ ASGITransport)."""
    async with _async_client_for_fastapi_app() as client:
        yield client


@pytest.fixture(scope="session")
def openapi_schema():
    """Fixture pour le schéma OpenAPI (généré dynamiquement, pas depuis fichier)."""
    return app.openapi()


# --- Activité / statuts utilisateurs : Redis mémoire + comptes de test reproductibles ---
_ACTIVITY_REDIS_TEST_FILES = frozenset(
    {"test_activity_ping.py", "test_user_statuses.py"},
)


def _activity_status_test_basename(request) -> str:
    try:
        path = getattr(request.node, "path", None) or getattr(request.node, "fspath", None)
        return path.name if path is not None else ""
    except Exception:
        return ""


def _ensure_seed_user_for_activity_tests(
    db_session: Session, *, username: str, role: UserRole, legacy_tag: str
) -> None:
    """Comptes stables attendus par les tests ping / statuts (plus de seed implicite Docker)."""
    password_plain = "Test1234!"
    u = db_session.query(User).filter(User.username == username).first()
    hashed = hash_password(password_plain)
    if u is None:
        u = User(
            username=username,
            hashed_password=hashed,
            role=role,
            status=UserStatus.ACTIVE,
            is_active=True,
            legacy_external_contact_id=legacy_tag,
        )
        db_session.add(u)
    else:
        u.hashed_password = hashed
        u.role = role
        u.status = UserStatus.ACTIVE
        u.is_active = True
    db_session.commit()
    db_session.refresh(u)


@pytest.fixture(autouse=True)
def _activity_user_status_tests_env(request, monkeypatch, db_session):
    """
    Pour ``test_activity_ping`` et ``test_user_statuses`` :
    - remplace le client Redis global par une implémentation en mémoire ;
    - garantit ``superadmintest1`` / ``admintest1`` avec mot de passe ``Test1234!`` ;
    - réinitialise le cache de seuil ``ActivityService`` (évite la fuite entre tests).
    """
    if _activity_status_test_basename(request) not in _ACTIVITY_REDIS_TEST_FILES:
        yield
        return
    if request.node.get_closest_marker("no_db"):
        yield
        return

    from recyclic_api.core import redis as redis_core
    from recyclic_api.services.activity_service import (
        ActivityService,
        DEFAULT_ACTIVITY_THRESHOLD_MINUTES,
    )

    monkeypatch.setattr(redis_core, "redis_client", MemoryRedisForTests())
    ActivityService._cached_threshold_minutes = DEFAULT_ACTIVITY_THRESHOLD_MINUTES
    ActivityService._cache_expiration_timestamp = 0.0

    _ensure_seed_user_for_activity_tests(
        db_session,
        username="superadmintest1",
        role=UserRole.SUPER_ADMIN,
        legacy_tag="seed-sa-activity",
    )
    _ensure_seed_user_for_activity_tests(
        db_session,
        username="admintest1",
        role=UserRole.ADMIN,
        legacy_tag="seed-ad-activity",
    )

    yield

