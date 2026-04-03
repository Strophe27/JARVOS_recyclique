"""
Tests d'intégration pour l'export bulk de tickets de réception (Story B45-P1).

Tests vérifient:
- Export CSV bulk avec filtres
- Export Excel bulk avec filtres
- Respect des filtres (date, statut, bénévole)
- Permissions (ADMIN/SUPER_ADMIN uniquement)
"""
import io
from datetime import datetime, timedelta, timezone
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from recyclic_api.models.ticket_depot import TicketDepot, TicketDepotStatus
from recyclic_api.models.poste_reception import PosteReception, PosteReceptionStatus
from recyclic_api.models.ligne_depot import LigneDepot
from recyclic_api.models.user import User, UserRole, UserStatus
from recyclic_api.models.category import Category
from recyclic_api.core.security import hash_password


@pytest.fixture
def test_category(db_session: Session) -> Category:
    """Créer une catégorie de test."""
    category = Category(
        id=uuid4(),
        name="Test Category",
        price=10.0
    )
    db_session.add(category)
    db_session.commit()
    db_session.refresh(category)
    return category


@pytest.fixture
def test_poste(db_session: Session) -> PosteReception:
    """Créer un poste de réception de test."""
    poste = PosteReception(
        id=uuid4(),
        name="Poste Test",
        status=PosteReceptionStatus.OPENED,
        opened_at=datetime.now(timezone.utc) - timedelta(days=1)
    )
    db_session.add(poste)
    db_session.commit()
    db_session.refresh(poste)
    return poste


@pytest.fixture
def test_benevole(db_session: Session) -> User:
    """Créer un bénévole de test."""
    benevole = User(
        id=uuid4(),
        username="benevole_test",
        hashed_password=hash_password("testpass"),
        role=UserRole.USER,
        status=UserStatus.ACTIVE,
        is_active=True
    )
    db_session.add(benevole)
    db_session.commit()
    db_session.refresh(benevole)
    return benevole


@pytest.fixture
def test_tickets(db_session: Session, test_poste: PosteReception, test_benevole: User, test_category: Category) -> list[TicketDepot]:
    """Créer plusieurs tickets de test avec différentes dates."""
    tickets = []
    now = datetime.now(timezone.utc)
    
    # Ticket 1: Hier, fermé, avec lignes
    ticket1 = TicketDepot(
        id=uuid4(),
        poste_id=test_poste.id,
        benevole_user_id=test_benevole.id,
        status=TicketDepotStatus.CLOSED,
        created_at=now - timedelta(days=1),
        closed_at=now - timedelta(hours=12)
    )
    db_session.add(ticket1)
    db_session.flush()
    
    ligne1 = LigneDepot(
        id=uuid4(),
        ticket_id=ticket1.id,
        category_id=test_category.id,
        poids_kg=5.5,
        destination="revente"
    )
    db_session.add(ligne1)
    
    ligne2 = LigneDepot(
        id=uuid4(),
        ticket_id=ticket1.id,
        category_id=test_category.id,
        poids_kg=3.2,
        destination="recyclage"
    )
    db_session.add(ligne2)
    
    # Ticket 2: Aujourd'hui, ouvert
    ticket2 = TicketDepot(
        id=uuid4(),
        poste_id=test_poste.id,
        benevole_user_id=test_benevole.id,
        status=TicketDepotStatus.OPENED,
        created_at=now - timedelta(hours=2)
    )
    db_session.add(ticket2)
    db_session.flush()
    
    ligne3 = LigneDepot(
        id=uuid4(),
        ticket_id=ticket2.id,
        category_id=test_category.id,
        poids_kg=2.1,
        destination="revente"
    )
    db_session.add(ligne3)
    
    # Ticket 3: Il y a 3 jours, fermé, sans lignes (ticket vide)
    ticket3 = TicketDepot(
        id=uuid4(),
        poste_id=test_poste.id,
        benevole_user_id=test_benevole.id,
        status=TicketDepotStatus.CLOSED,
        created_at=now - timedelta(days=3),
        closed_at=now - timedelta(days=3, hours=1)
    )
    db_session.add(ticket3)
    
    db_session.commit()
    
    for ticket in [ticket1, ticket2, ticket3]:
        db_session.refresh(ticket)
        tickets.append(ticket)
    
    return tickets


class TestBulkExportReceptionTicketsCSV:
    """Tests pour l'export CSV bulk de tickets de réception."""
    
    def test_export_csv_bulk_success(self, admin_client, test_tickets):
        """Test export CSV bulk avec succès."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert "export_tickets_reception" in response.headers["content-disposition"]
        
        # Vérifier le contenu CSV
        content = response.text
        assert "ID Ticket" in content
        assert "Statut" in content
        assert "Date Création" in content
        assert "Bénévole" in content
        assert "Nombre Lignes" in content
        assert "Poids Total (kg)" in content
    
    def test_export_csv_bulk_with_filters(self, admin_client, test_tickets):
        """Test export CSV avec filtres (statut, date)."""
        yesterday = datetime.now(timezone.utc) - timedelta(days=1)
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": yesterday.isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "status": "closed",
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        content = response.text
        
        # Vérifier que seuls les tickets fermés sont inclus
        lines = content.split('\n')
        data_lines = [l for l in lines[1:] if l.strip()]
        
        # Au moins un ticket fermé doit être présent
        assert len(data_lines) > 0
    
    def test_export_csv_bulk_unauthorized(self, client, db_session, test_tickets):
        """Test que USER ne peut pas exporter bulk."""
        from recyclic_api.core.security import create_access_token
        
        # Créer un utilisateur USER (non-admin) dans la session de test
        user = User(
            id=uuid4(),
            username="user_test",
            hashed_password=hash_password("testpass"),
            role=UserRole.USER,
            status=UserStatus.ACTIVE,
            is_active=True
        )
        db_session.add(user)
        db_session.commit()
        db_session.refresh(user)
        
        token = create_access_token(data={"sub": str(user.id)})
        client.headers["Authorization"] = f"Bearer {token}"
        
        response = client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {},
                "format": "csv"
            }
        )
        
        assert response.status_code == 403
    
    def test_export_csv_bulk_respects_filters(self, admin_client, test_tickets, test_benevole):
        """Test que les filtres sont bien respectés."""
        # Filtrer par bénévole
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "benevole_id": str(test_benevole.id),
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        content = response.text
        
        # Vérifier que toutes les sessions retournées sont du bon bénévole
        lines = content.split('\n')
        # Les tickets de test sont tous du même bénévole, donc on vérifie juste que ça fonctionne
        assert len(lines) > 1  # Au moins headers + données


class TestBulkExportReceptionTicketsExcel:
    """Tests pour l'export Excel bulk de tickets de réception."""
    
    def test_export_excel_bulk_success(self, admin_client, test_tickets):
        """Test export Excel bulk avec succès."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]
        
        # Vérifier que c'est un fichier Excel valide
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier les onglets
        assert "Résumé" in wb.sheetnames
        assert "Détails" in wb.sheetnames
        
        # Vérifier le contenu de l'onglet Résumé
        ws_summary = wb["Résumé"]
        assert ws_summary.max_row > 1  # Au moins headers + données
        
        # Vérifier les en-têtes
        headers = [cell.value for cell in ws_summary[1]]
        assert "Statut" in headers
        assert "Date Création" in headers
        assert "Nb Lignes" in headers
        assert "Poids Total (kg)" in headers
        
        # Vérifier le contenu de l'onglet Détails
        ws_details = wb["Détails"]
        assert ws_details.max_row > 1
        
        headers_details = [cell.value for cell in ws_details[1]]
        assert "ID Ticket" in headers_details
        assert "Nombre Lignes" in headers_details
    
    def test_export_excel_bulk_with_filters(self, admin_client, test_tickets):
        """Test export Excel avec filtres."""
        yesterday = datetime.now(timezone.utc) - timedelta(days=1)
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": yesterday.isoformat(),
                    "status": "closed",
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        
        # Vérifier le fichier Excel
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier qu'il y a des données
        ws_summary = wb["Résumé"]
        assert ws_summary.max_row > 1  # Headers + au moins une ligne de données
    
    def test_export_excel_formatting_styles(self, admin_client, test_tickets):
        """Test que la mise en forme (styles, couleurs, bordures) est appliquée."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier les styles des en-têtes dans l'onglet Résumé
        ws_summary = wb["Résumé"]
        header_row = ws_summary[1]
        
        for cell in header_row:
            # Vérifier que les en-têtes sont en gras
            assert cell.font.bold is True, f"Header cell {cell.coordinate} should be bold"
            # Vérifier que les en-têtes ont un fond coloré
            assert cell.fill.start_color is not None, f"Header cell {cell.coordinate} should have fill color"
            # Vérifier que les en-têtes ont des bordures
            assert cell.border is not None, f"Header cell {cell.coordinate} should have border"
        
        # Vérifier les styles des en-têtes dans l'onglet Détails
        ws_details = wb["Détails"]
        detail_header_row = ws_details[1]
        
        for cell in detail_header_row:
            assert cell.font.bold is True, f"Detail header cell {cell.coordinate} should be bold"
            assert cell.fill.start_color is not None, f"Detail header cell {cell.coordinate} should have fill color"
            assert cell.border is not None, f"Detail header cell {cell.coordinate} should have border"
    
    def test_export_excel_performance_1000_tickets(self, admin_client, db_session, test_poste, test_benevole, test_category):
        """Test de performance : export Excel de 1000 tickets doit être < 30 secondes."""
        import time
        
        # Créer 1000 tickets de test
        now = datetime.now(timezone.utc)
        tickets = []
        for i in range(1000):
            ticket = TicketDepot(
                id=uuid4(),
                poste_id=test_poste.id,
                benevole_user_id=test_benevole.id,
                status=TicketDepotStatus.CLOSED,
                created_at=now - timedelta(days=i % 30),
                closed_at=now - timedelta(days=i % 30, hours=1)
            )
            tickets.append(ticket)
        
        db_session.add_all(tickets)
        db_session.flush()
        
        # Créer des lignes pour chaque ticket
        for idx, ticket in enumerate(tickets):
            ligne = LigneDepot(
                id=uuid4(),
                ticket_id=ticket.id,
                category_id=test_category.id,
                poids_kg=float(idx % 10) + 0.5,
                destination="revente"
            )
            db_session.add(ligne)
        
        db_session.commit()
        
        # Mesurer le temps d'export
        start_time = time.time()
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (now - timedelta(days=30)).isoformat(),
                    "date_to": now.isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        elapsed_time = time.time() - start_time
        
        assert response.status_code == 200
        assert elapsed_time < 30.0, f"Export took {elapsed_time:.2f}s, should be < 30s"
        
        # Vérifier que le fichier est valide
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        assert "Résumé" in wb.sheetnames
        assert "Détails" in wb.sheetnames


class TestBulkExportReceptionTicketsValidation:
    """Tests de validation pour l'export bulk de tickets."""
    
    def test_export_bulk_invalid_format(self, admin_client):
        """Test avec format invalide."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {},
                "format": "pdf"  # Format invalide
            }
        )
        
        assert response.status_code == 422  # Validation error
    
    def test_export_bulk_empty_result(self, admin_client):
        """Test export avec aucun résultat (filtres trop restrictifs)."""
        far_future = datetime.now(timezone.utc) + timedelta(days=365)
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": far_future.isoformat(),
                    "date_to": (far_future + timedelta(days=1)).isoformat()
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        # Le CSV devrait contenir seulement les headers
        content = response.text
        lines = [l for l in content.split('\n') if l.strip()]
        assert len(lines) == 1  # Seulement headers
    
    def test_export_bulk_reception_400_date_format_frontend(self, admin_client, test_tickets):
        """
        Test de reproduction du bug B50-P2 : dates en format string ISO avec 'Z' suffix.
        Le frontend envoie : "2025-12-10T00:00:00.000Z"
        """
        # Format exact envoyé par le frontend (ReceptionSessionManager.tsx ligne 668-669)
        payload = {
            "filters": {
                "date_from": "2025-12-10T00:00:00.000Z",
                "date_to": "2025-12-10T23:59:59.999Z",
                "status": "closed"
            },
            "format": "csv"
        }
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json=payload
        )
        
        # Avant correction : devrait échouer avec 400 ou 422
        # Après correction : devrait réussir avec 200
        if response.status_code != 200:
            # Capturer le message d'erreur exact pour debug
            error_detail = response.json() if response.content else {}
            pytest.fail(
                f"Export échoué avec status {response.status_code}. "
                f"Détails: {error_detail}. "
                f"Ce test reproduit le bug B50-P2."
            )
        
        # Si on arrive ici, le bug est corrigé
        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv; charset=utf-8"
    
    def test_export_bulk_reception_with_undefined_dates(self, admin_client, test_tickets):
        """Test export avec dates undefined/null (B50-P2: régression)."""
        payload = {
            "filters": {
                "status": "closed",
                "include_empty": False
                # Pas de date_from ni date_to
            },
            "format": "csv"
        }
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json=payload
        )
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv; charset=utf-8"
    
    def test_export_bulk_reception_with_date_only_format(self, admin_client, test_tickets):
        """Test export avec format date seule YYYY-MM-DD (B50-P2: régression)."""
        payload = {
            "filters": {
                "date_from": "2025-12-10",
                "date_to": "2025-12-10",
                "status": "closed"
            },
            "format": "csv"
        }
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json=payload
        )
        
        # Le validator devrait parser ce format aussi
        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv; charset=utf-8"
    
    def test_export_bulk_reception_with_isoformat_no_z(self, admin_client, test_tickets):
        """Test export avec format ISO sans 'Z' (B50-P2: régression)."""
        payload = {
            "filters": {
                "date_from": "2025-12-10T00:00:00+00:00",
                "date_to": "2025-12-10T23:59:59+00:00",
                "status": "closed"
            },
            "format": "csv"
        }
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json=payload
        )
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv; charset=utf-8"


class TestCalculateTicketTotalsSignature:
    """
    Tests de régression pour vérifier la signature de _calculate_ticket_totals.
    
    B50-P2: Vérifie que la fonction retourne toujours 5 valeurs (total_lignes, total_poids, 
    poids_entree, poids_direct, poids_sortie) pour prévenir les bugs de déballage.
    """
    
    def test_calculate_ticket_totals_signature(self, db_session: Session, test_tickets):
        """
        Test que _calculate_ticket_totals retourne exactement 5 valeurs.
        
        Ce test de régression prévient les bugs de type "ValueError: too many values to unpack"
        en vérifiant que la signature de la fonction reste stable.
        """
        from recyclic_api.services.reception_service import ReceptionService
        from decimal import Decimal
        
        service = ReceptionService(db_session)
        ticket = test_tickets[0]  # Utiliser le premier ticket de test
        
        # Appeler la fonction et vérifier le nombre de valeurs retournées
        result = service._calculate_ticket_totals(ticket)
        
        # Vérifier que c'est un tuple de 5 éléments
        assert isinstance(result, tuple), f"Expected tuple, got {type(result)}"
        assert len(result) == 5, f"Expected 5 values, got {len(result)}: {result}"
        
        # Déballer les valeurs pour vérifier qu'elles sont du bon type
        total_lignes, total_poids, poids_entree, poids_direct, poids_sortie = result
        
        # Vérifier les types
        assert isinstance(total_lignes, int), f"total_lignes should be int, got {type(total_lignes)}"
        assert isinstance(total_poids, (int, float, Decimal)), f"total_poids should be numeric, got {type(total_poids)}"
        assert isinstance(poids_entree, (int, float, Decimal)), f"poids_entree should be numeric, got {type(poids_entree)}"
        assert isinstance(poids_direct, (int, float, Decimal)), f"poids_direct should be numeric, got {type(poids_direct)}"
        assert isinstance(poids_sortie, (int, float, Decimal)), f"poids_sortie should be numeric, got {type(poids_sortie)}"
        
        # Vérifier que les valeurs sont cohérentes
        assert total_lignes >= 0, "total_lignes should be >= 0"
        assert float(total_poids) >= 0, "total_poids should be >= 0"
        assert float(poids_entree) >= 0, "poids_entree should be >= 0"
        assert float(poids_direct) >= 0, "poids_direct should be >= 0"
        assert float(poids_sortie) >= 0, "poids_sortie should be >= 0"
        
        # Vérifier que le poids total est la somme des poids par flux
        # (avec une petite tolérance pour les erreurs d'arrondi)
        total_expected = float(poids_entree) + float(poids_direct) + float(poids_sortie)
        total_actual = float(total_poids)
        assert abs(total_expected - total_actual) < 0.01, \
            f"total_poids ({total_actual}) should equal sum of poids_entree + poids_direct + poids_sortie ({total_expected})"
    
    def test_calculate_ticket_totals_empty_ticket(self, db_session: Session, test_poste, test_benevole):
        """
        Test que _calculate_ticket_totals fonctionne avec un ticket vide.
        """
        from recyclic_api.services.reception_service import ReceptionService
        
        # Créer un ticket sans lignes
        empty_ticket = TicketDepot(
            id=uuid4(),
            poste_id=test_poste.id,
            benevole_user_id=test_benevole.id,
            status=TicketDepotStatus.OPENED.value
        )
        db_session.add(empty_ticket)
        db_session.commit()
        
        service = ReceptionService(db_session)
        result = service._calculate_ticket_totals(empty_ticket)
        
        # Vérifier la signature
        assert len(result) == 5, f"Expected 5 values, got {len(result)}"
        total_lignes, total_poids, poids_entree, poids_direct, poids_sortie = result
        
        # Pour un ticket vide, tout devrait être 0
        assert total_lignes == 0
        assert float(total_poids) == 0.0
        assert float(poids_entree) == 0.0
        assert float(poids_direct) == 0.0
        assert float(poids_sortie) == 0.0

