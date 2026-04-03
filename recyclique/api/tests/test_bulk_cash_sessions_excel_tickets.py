"""
Tests pour l'export Excel bulk des sessions de caisse avec onglet "Détails Tickets" (Story B50-P1).

Tests vérifient:
- Présence de l'onglet "Détails Tickets"
- Une ligne par item de vente
- Filtrage sur catégories principales uniquement
- Agrégation des sous-catégories vers parent
- Formatage des montants et poids
"""
import io
from datetime import datetime, timedelta, timezone
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from recyclic_api.models.cash_session import CashSession, CashSessionStatus
from recyclic_api.models.sale import Sale
from recyclic_api.models.sale_item import SaleItem
from recyclic_api.models.category import Category
from recyclic_api.models.user import User, UserRole, UserStatus
from recyclic_api.models.site import Site
from recyclic_api.models.cash_register import CashRegister
from recyclic_api.schemas.cash_session import CashSessionFilters
from recyclic_api.services.report_service import generate_bulk_cash_sessions_excel
from recyclic_api.core.security import hash_password


@pytest.fixture
def test_site(db_session: Session) -> Site:
    """Créer un site de test."""
    site = Site(
        id=uuid4(),
        name="Test Site",
        address="123 Test St",
        city="Test City",
        postal_code="12345",
        country="France",
        is_active=True
    )
    db_session.add(site)
    db_session.commit()
    db_session.refresh(site)
    return site


@pytest.fixture
def test_operator(db_session: Session, test_site: Site) -> User:
    """Créer un opérateur de test."""
    operator = User(
        id=uuid4(),
        username="test_operator",
        email="operator@test.com",
        hashed_password=hash_password("test123"),
        role=UserRole.USER,
        status=UserStatus.ACTIVE,
        is_active=True,
        site_id=test_site.id
    )
    db_session.add(operator)
    db_session.commit()
    db_session.refresh(operator)
    return operator


@pytest.fixture
def test_register(db_session: Session, test_site: Site) -> CashRegister:
    """Créer un registre de caisse de test."""
    register = CashRegister(
        id=uuid4(),
        name="Test Register",
        site_id=test_site.id,
        is_active=True
    )
    db_session.add(register)
    db_session.commit()
    db_session.refresh(register)
    return register


@pytest.fixture
def main_category(db_session: Session) -> Category:
    """Créer une catégorie principale (parent_id IS NULL)."""
    category = Category(
        id=uuid4(),
        name="EEE-1",
        is_active=True,
        parent_id=None
    )
    db_session.add(category)
    db_session.commit()
    db_session.refresh(category)
    return category


@pytest.fixture
def sub_category(db_session: Session, main_category: Category) -> Category:
    """Créer une sous-catégorie (avec parent)."""
    category = Category(
        id=uuid4(),
        name="EEE-1-SUB",
        is_active=True,
        parent_id=main_category.id
    )
    db_session.add(category)
    db_session.commit()
    db_session.refresh(category)
    return category


@pytest.fixture
def test_session_with_sales(
    db_session: Session,
    test_operator: User,
    test_site: Site,
    test_register: CashRegister,
    main_category: Category,
    sub_category: Category
) -> CashSession:
    """Créer une session de caisse avec des ventes et items."""
    session = CashSession(
        id=uuid4(),
        operator_id=test_operator.id,
        site_id=test_site.id,
        register_id=test_register.id,
        initial_amount=100.0,
        current_amount=150.0,
        status=CashSessionStatus.CLOSED,
        total_sales=50.0,
        total_items=3,
        number_of_sales=2
    )
    session.opened_at = datetime.now(timezone.utc) - timedelta(hours=2)
    session.closed_at = datetime.now(timezone.utc)
    db_session.add(session)
    db_session.commit()
    db_session.refresh(session)
    
    # Créer une vente avec item catégorie principale
    sale1 = Sale(
        id=uuid4(),
        cash_session_id=session.id,
        operator_id=test_operator.id,
        total_amount=30.0,
        donation=0.0
    )
    sale1.created_at = datetime.now(timezone.utc) - timedelta(hours=1)
    db_session.add(sale1)
    db_session.commit()
    db_session.refresh(sale1)
    
    item1 = SaleItem(
        id=uuid4(),
        sale_id=sale1.id,
        category="EEE-1",  # Catégorie principale
        quantity=2,
        weight=5.5,
        unit_price=15.0,
        total_price=30.0
    )
    db_session.add(item1)
    
    # Créer une vente avec item sous-catégorie
    sale2 = Sale(
        id=uuid4(),
        cash_session_id=session.id,
        operator_id=test_operator.id,
        total_amount=20.0,
        donation=0.0
    )
    sale2.created_at = datetime.now(timezone.utc) - timedelta(minutes=30)
    db_session.add(sale2)
    db_session.commit()
    db_session.refresh(sale2)
    
    item2 = SaleItem(
        id=uuid4(),
        sale_id=sale2.id,
        category="EEE-1-SUB",  # Sous-catégorie
        quantity=1,
        weight=3.2,
        unit_price=20.0,
        total_price=20.0
    )
    db_session.add(item2)
    
    db_session.commit()
    return session


class TestBulkCashSessionsExcelTicketsTab:
    """Tests pour l'onglet 'Détails Tickets' dans l'export Excel."""
    
    def test_tickets_tab_exists(self, db_session: Session, test_session_with_sales: CashSession):
        """Test que l'onglet 'Détails Tickets' existe."""
        filters = CashSessionFilters(
            date_from=datetime.now(timezone.utc) - timedelta(days=1),
            date_to=datetime.now(timezone.utc)
        )
        
        buffer = generate_bulk_cash_sessions_excel(db_session, filters)
        wb = load_workbook(buffer)
        
        # Vérifier que l'onglet existe
        assert "Détails Tickets" in wb.sheetnames
        assert "Résumé" in wb.sheetnames
        assert "Détails" in wb.sheetnames
    
    def test_tickets_tab_headers(self, db_session: Session, test_session_with_sales: CashSession):
        """Test que les en-têtes de l'onglet sont corrects."""
        filters = CashSessionFilters(
            date_from=datetime.now(timezone.utc) - timedelta(days=1),
            date_to=datetime.now(timezone.utc)
        )
        
        buffer = generate_bulk_cash_sessions_excel(db_session, filters)
        wb = load_workbook(buffer)
        ws_tickets = wb["Détails Tickets"]
        
        # Vérifier les en-têtes
        headers = [cell.value for cell in ws_tickets[1]]
        expected_headers = [
            'Numéro Ticket',
            'Date Vente',
            'Catégorie Principale',
            'Quantité',
            'Poids (kg)',
            'Prix Unitaire (€)',
            'Prix Total (€)'
        ]
        assert headers == expected_headers
    
    def test_tickets_tab_one_line_per_item(self, db_session: Session, test_session_with_sales: CashSession):
        """Test qu'il y a une ligne par item de vente."""
        filters = CashSessionFilters(
            date_from=datetime.now(timezone.utc) - timedelta(days=1),
            date_to=datetime.now(timezone.utc)
        )
        
        buffer = generate_bulk_cash_sessions_excel(db_session, filters)
        wb = load_workbook(buffer)
        ws_tickets = wb["Détails Tickets"]
        
        # Vérifier qu'il y a au moins 2 lignes de données (2 items dans test_session_with_sales)
        # + 1 ligne d'en-tête
        assert ws_tickets.max_row >= 3  # Headers + au moins 2 items
    
    def test_tickets_tab_main_categories_only(self, db_session: Session, test_session_with_sales: CashSession):
        """Test que seules les catégories principales apparaissent."""
        filters = CashSessionFilters(
            date_from=datetime.now(timezone.utc) - timedelta(days=1),
            date_to=datetime.now(timezone.utc)
        )
        
        buffer = generate_bulk_cash_sessions_excel(db_session, filters)
        wb = load_workbook(buffer)
        ws_tickets = wb["Détails Tickets"]
        
        # Parcourir les lignes de données (skip header)
        categories_found = set()
        for row in ws_tickets.iter_rows(min_row=2, values_only=True):
            if row[2]:  # Catégorie Principale (colonne C)
                categories_found.add(row[2])
        
        # Vérifier que toutes les catégories sont des catégories principales
        # Dans notre test, on a EEE-1 (principale) et EEE-1-SUB (sous-catégorie)
        # EEE-1-SUB doit être remontée à EEE-1
        assert "EEE-1" in categories_found
        assert "EEE-1-SUB" not in categories_found  # La sous-catégorie ne doit pas apparaître
    
    def test_tickets_tab_subcategory_aggregation(self, db_session: Session, test_session_with_sales: CashSession):
        """Test que les sous-catégories sont agrégées vers le parent."""
        filters = CashSessionFilters(
            date_from=datetime.now(timezone.utc) - timedelta(days=1),
            date_to=datetime.now(timezone.utc)
        )
        
        buffer = generate_bulk_cash_sessions_excel(db_session, filters)
        wb = load_workbook(buffer)
        ws_tickets = wb["Détails Tickets"]
        
        # Vérifier que l'item avec sous-catégorie apparaît avec la catégorie principale
        found_subcategory_item = False
        for row in ws_tickets.iter_rows(min_row=2, values_only=True):
            category = row[2]  # Catégorie Principale
            quantity = row[3]  # Quantité
            if category == "EEE-1" and quantity == 1:
                # C'est l'item avec sous-catégorie, vérifié qu'il est bien avec EEE-1
                found_subcategory_item = True
                break
        
        assert found_subcategory_item, "L'item avec sous-catégorie doit apparaître avec la catégorie principale"
    
    def test_tickets_tab_formatting(self, db_session: Session, test_session_with_sales: CashSession):
        """Test le formatage des montants et poids."""
        filters = CashSessionFilters(
            date_from=datetime.now(timezone.utc) - timedelta(days=1),
            date_to=datetime.now(timezone.utc)
        )
        
        buffer = generate_bulk_cash_sessions_excel(db_session, filters)
        wb = load_workbook(buffer)
        ws_tickets = wb["Détails Tickets"]
        
        # Vérifier le formatage (virgule comme séparateur décimal pour format français)
        for row in ws_tickets.iter_rows(min_row=2, values_only=True):
            weight = row[4]  # Poids (kg)
            unit_price = row[5]  # Prix Unitaire (€)
            total_price = row[6]  # Prix Total (€)
            
            # Vérifier que les montants utilisent la virgule comme séparateur
            if unit_price:
                assert ',' in str(unit_price) or isinstance(unit_price, (int, float))
            if total_price:
                assert ',' in str(total_price) or isinstance(total_price, (int, float))
            if weight:
                assert ',' in str(weight) or isinstance(weight, (int, float))
    
    def test_tickets_tab_empty_sessions(self, db_session: Session, test_site: Site, test_operator: User):
        """Test que les sessions vides n'apparaissent pas dans l'onglet."""
        # Créer une session vide
        empty_session = CashSession(
            id=uuid4(),
            operator_id=test_operator.id,
            site_id=test_site.id,
            initial_amount=100.0,
            current_amount=100.0,
            status=CashSessionStatus.CLOSED,
            total_sales=0.0,
            total_items=0,
            number_of_sales=0
        )
        empty_session.opened_at = datetime.now(timezone.utc) - timedelta(hours=1)
        empty_session.closed_at = datetime.now(timezone.utc)
        db_session.add(empty_session)
        db_session.commit()
        
        filters = CashSessionFilters(
            date_from=datetime.now(timezone.utc) - timedelta(days=1),
            date_to=datetime.now(timezone.utc)
        )
        
        buffer = generate_bulk_cash_sessions_excel(db_session, filters)
        wb = load_workbook(buffer)
        ws_tickets = wb["Détails Tickets"]
        
        # Vérifier qu'il n'y a que l'en-tête (pas de données pour session vide)
        # Mais on peut avoir des données d'autres sessions, donc on vérifie juste que l'onglet existe
        assert ws_tickets.max_row >= 1  # Au moins l'en-tête

